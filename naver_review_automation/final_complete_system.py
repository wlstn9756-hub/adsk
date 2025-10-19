from fastapi import FastAPI, Depends, HTTPException, Form, Request, Response, File, UploadFile, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Date, Text, Float, Boolean, ForeignKey, func, or_
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship, joinedload
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date, timedelta
import hashlib
import secrets
import pandas as pd
from io import BytesIO, StringIO
import os
import json
import requests
from bs4 import BeautifulSoup
import re
import csv
import openpyxl
import asyncio
import threading
from concurrent.futures import ThreadPoolExecutor

# 데이터베이스 설정
# 데이터베이스 파일 절대 경로 설정
BASE_DB_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = f"sqlite:///{os.path.join(BASE_DB_DIR, 'final_complete_system.db')}"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# FastAPI 앱 생성
app = FastAPI()

# HTTPException 핸들러 추가 (API 엔드포인트에서 JSON 응답 보장)
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """API 엔드포인트에서 HTTPException 발생 시 항상 JSON 응답 반환"""
    # API 경로인 경우 JSON 응답 반환
    if request.url.path.startswith("/api/"):
        return JSONResponse(
            status_code=exc.status_code,
            content={"success": False, "message": exc.detail}
        )
    # 일반 경로인 경우 로그인 페이지로 리다이렉트 (401 에러)
    if exc.status_code == 401:
        return RedirectResponse(url="/login", status_code=302)
    # 그 외 에러는 JSON 응답
    return JSONResponse(
        status_code=exc.status_code,
        content={"success": False, "message": exc.detail}
    )

# 템플릿 디렉토리 절대 경로 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# 세션 저장소 (메모리 기반)
sessions = {}

# 데이터베이스 모델
class Company(Base):
    __tablename__ = "companies"

    id = Column(Integer, primary_key=True)
    name = Column(String(100), unique=True, nullable=False)
    display_name = Column(String(100), nullable=False)
    unit_price = Column(Float, default=3000)  # 고객사별 단가
    manager_id = Column(Integer, ForeignKey('managers.id'), nullable=True)  # 담당자 연결
    created_at = Column(DateTime, default=datetime.now)

    users = relationship("User", back_populates="company")
    orders = relationship("ReceiptWorkOrder", back_populates="company")
    manager = relationship("Manager")

class User(Base):
    __tablename__ = "users"

    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True, nullable=False)
    password_hash = Column(String(128), nullable=False)
    full_name = Column(String(100))
    email = Column(String(100))
    phone = Column(String(20))
    role = Column(String(20), default="reviewer")  # super_admin, company_admin, reviewer
    company_id = Column(Integer, ForeignKey("companies.id"))
    commission_rate = Column(Float, default=0.1)  # 인센티브 비율 (10%)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.now)

    company = relationship("Company", back_populates="users")
    orders = relationship("ReceiptWorkOrder", back_populates="client", foreign_keys="ReceiptWorkOrder.client_id")
    extension_requests = relationship("ExtensionRequest", foreign_keys="ExtensionRequest.client_id", back_populates="client")

class ExtensionRequest(Base):
    __tablename__ = "extension_requests"

    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(Integer, ForeignKey('receipt_work_orders.id'), nullable=False)
    client_id = Column(Integer, ForeignKey('users.id'), nullable=False)
    daily_count = Column(Integer, nullable=False)
    total_days = Column(Integer, nullable=False)
    status = Column(String(20), default='pending')  # pending, approved, rejected
    created_at = Column(DateTime, default=datetime.now)
    processed_at = Column(DateTime, nullable=True)
    processed_by = Column(Integer, ForeignKey('users.id'), nullable=True)

    # 관계
    order = relationship("ReceiptWorkOrder", back_populates="extension_requests")
    client = relationship("User", foreign_keys=[client_id], back_populates="extension_requests")
    processor = relationship("User", foreign_keys=[processed_by])

# 담당자 모델 (로그인 불필요, 이름과 인센티브 비율만)
class Manager(Base):
    __tablename__ = 'managers'

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    commission_rate = Column(Float, default=0.1)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.now)

class ReceiptWorkOrder(Base):
    __tablename__ = "receipt_work_orders"

    id = Column(Integer, primary_key=True)
    order_no = Column(String(50), unique=True, nullable=False)
    company_id = Column(Integer, ForeignKey("companies.id"), nullable=False)
    client_id = Column(Integer, ForeignKey("users.id"), nullable=False)

    # 업체 정보
    business_name = Column(String(200), nullable=False)
    representative_name = Column(String(100), nullable=False)
    business_number = Column(String(20), nullable=False)
    business_type = Column(String(20), nullable=False, default="일반")  # 업종 (맛집/일반)
    place_number = Column(String(20))
    place_link = Column(String(500))  # 플레이스 링크 추가
    business_address = Column(Text, nullable=False)

    # 작업 정보
    receipt_date = Column(Date, nullable=False)
    start_date = Column(Date, nullable=False)
    end_date = Column(Date, nullable=False)
    daily_count = Column(Integer, nullable=False)
    working_days = Column(Integer, nullable=False)
    total_count = Column(Integer, nullable=False)
    completed_count = Column(Integer, default=0)

    # 가이드라인
    guidelines = Column(Text)

    # 가격 정보
    unit_price = Column(Float, default=3000)
    total_price = Column(Float, nullable=False)

    # 첨부 이미지 (영수증/사업자등록증)
    attachment_images = Column(Text)  # JSON 배열 형식으로 저장: ["path1.jpg", "path2.jpg"]

    # 리뷰 자료 (복붙멘트 & 사진)
    review_excel_path = Column(Text)  # 엑셀 파일 경로
    review_photos_path = Column(Text)  # 사진 ZIP 파일 경로

    # 상태 정보
    status = Column(String(20), default="pending")  # pending, approved, rejected, in_progress, completed
    rejection_reason = Column(Text)
    approved_at = Column(DateTime)
    approved_by = Column(Integer, ForeignKey("users.id"))
    completed_at = Column(DateTime)

    # 관리자 메모
    admin_memo = Column(Text)

    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)

    # 관계
    company = relationship("Company", back_populates="orders")
    client = relationship("User", back_populates="orders", foreign_keys=[client_id])
    reviews = relationship("Review", back_populates="order")
    extension_requests = relationship("ExtensionRequest", back_populates="order")

class Review(Base):
    __tablename__ = "reviews"

    id = Column(Integer, primary_key=True)
    order_id = Column(Integer, ForeignKey("receipt_work_orders.id"), nullable=False)
    reviewer_id = Column(Integer, ForeignKey("users.id"))

    # 리뷰 URL 및 추출 정보
    review_url = Column(String(500))
    naver_review_id = Column(String(100))

    # 리뷰 내용
    content = Column(Text, nullable=False)
    rating = Column(Integer, default=5)
    review_date = Column(Date, nullable=False)
    receipt_number = Column(String(50))

    # 추출된 정보
    author_name = Column(String(100))
    visit_date = Column(String(50))
    receipt_date_str = Column(String(50))  # 영수증 날짜 문자열
    menu_items = Column(Text)  # JSON 형태로 메뉴 저장
    images = Column(Text)  # JSON 형태로 이미지 URL 저장

    # 상태 정보
    is_submitted = Column(Boolean, default=False)
    submitted_at = Column(DateTime)
    extracted_at = Column(DateTime)

    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)

    order = relationship("ReceiptWorkOrder", back_populates="reviews")

# 데이터베이스 연결
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# 비밀번호 해시 함수
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

# 세션 관리 함수
def create_session(user_id: int) -> str:
    session_id = secrets.token_urlsafe(32)
    sessions[session_id] = {
        "user_id": user_id,
        "created_at": datetime.now()
    }
    return session_id

def get_current_user(request: Request, db: Session = Depends(get_db)):
    session_id = request.cookies.get("session_id")
    if not session_id or session_id not in sessions:
        return None

    session = sessions[session_id]
    user = db.query(User).filter(User.id == session["user_id"]).first()
    return user

def require_login(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다")
    return user

def require_admin(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    if user.role not in ["super_admin", "company_admin"]:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다")
    return user

def require_super_admin(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    if user.role != "super_admin":
        raise HTTPException(status_code=403, detail="시스템 관리자 권한이 필요합니다")
    return user

# 파일 읽기 함수 (CSV, Excel, 한컴 지원)
def read_file_to_dataframe(file_content: bytes, filename: str):
    """다양한 파일 형식을 DataFrame으로 변환"""
    try:
        # 파일 확장자 확인
        ext = filename.lower().split('.')[-1]
        print(f"[DEBUG] Processing file: {filename}, extension: {ext}, size: {len(file_content)} bytes")

        if ext in ['xlsx', 'xls']:
            # Excel 파일 - 여러 방법으로 시도
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning)

            # 방법 1: xlrd 엔진으로 시도 (구형 xls 파일)
            if ext == 'xls':
                try:
                    df = pd.read_excel(BytesIO(file_content), engine='xlrd')
                    print(f"[DEBUG] XLRD success: {len(df)} rows, columns: {list(df.columns)}")
                    return df
                except Exception as e:
                    print(f"[DEBUG] XLRD failed: {str(e)}")

            # 방법 2: openpyxl로 직접 읽기 (스타일 완전 무시)
            try:
                import openpyxl
                from openpyxl import load_workbook

                # 임시 파일로 저장 후 읽기 (스타일 오류 회피)
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(file_content)
                    tmp_path = tmp.name

                try:
                    wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_vba=False)
                    ws = wb.active

                    # 모든 데이터를 리스트로 추출
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            data.append(list(row))

                    if data and len(data) > 0:
                        # 첫 행을 컬럼으로 사용
                        df = pd.DataFrame(data[1:] if len(data) > 1 else [], columns=data[0] if data else [])
                        print(f"[DEBUG] Temp file method success: {len(df)} rows, columns: {list(df.columns)}")
                        import os
                        os.unlink(tmp_path)  # 임시 파일 삭제
                        return df
                finally:
                    import os
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
            except Exception as e:
                print(f"[DEBUG] Temp file method failed: {str(e)}")

            # 방법 3: pandas read_excel (calamine 엔진 시도)
            try:
                df = pd.read_excel(BytesIO(file_content), engine='calamine')
                print(f"[DEBUG] Calamine success: {len(df)} rows, columns: {list(df.columns)}")
                return df
            except Exception as e:
                print(f"[DEBUG] Calamine failed: {str(e)}")

            # 방법 4: 기본 pandas (엔진 자동 선택)
            try:
                df = pd.read_excel(BytesIO(file_content), engine=None)
                print(f"[DEBUG] Default pandas success: {len(df)} rows, columns: {list(df.columns)}")
                return df
            except Exception as e:
                print(f"[DEBUG] Default pandas failed: {str(e)}")

            # 모든 방법 실패
            print(f"[ERROR] All Excel reading methods failed for {filename}")
            return pd.DataFrame()
        elif ext == 'csv':
            # CSV 파일 - 기본 인코딩 시도
            try:
                text = file_content.decode('utf-8')
                print(f"[DEBUG] CSV decoded with UTF-8")
            except UnicodeDecodeError:
                try:
                    text = file_content.decode('cp949')
                    print(f"[DEBUG] CSV decoded with CP949")
                except UnicodeDecodeError:
                    text = file_content.decode('utf-8', errors='ignore')
                    print(f"[DEBUG] CSV decoded with UTF-8 (ignoring errors)")

            df = pd.read_csv(StringIO(text))
            print(f"[DEBUG] CSV read success: {len(df)} rows, columns: {list(df.columns)}")
            return df
        else:
            # 기본적으로 Excel 시도
            print(f"[DEBUG] Unknown extension {ext}, trying Excel")
            df = pd.read_excel(BytesIO(file_content))
            print(f"[DEBUG] Excel read success: {len(df)} rows, columns: {list(df.columns)}")
            return df
    except Exception as e:
        print(f"[ERROR] File reading failed: {str(e)}")
        raise ValueError(f"파일 읽기 실패: {str(e)}")

# 네이버 리뷰 추출 함수
def extract_naver_review(url: str):
    """네이버 리뷰 URL에서 리뷰 정보 추출"""
    try:
        # 네이버 리뷰 API 호출을 시뮬레이션 (실제로는 스크래핑 필요)
        review_data = {
            "author_name": "리뷰어",
            "content": "맛있게 잘 먹었습니다. 서비스도 친절하고 좋았어요.",
            "rating": 5,
            "visit_date": "2024년 1월 방문",
            "menu_items": ["아메리카노", "카페라떼"],
            "images": []
        }
        return review_data
    except Exception as e:
        print(f"Error extracting review: {str(e)}")
        return None

# API 엔드포인트
@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    user = get_current_user(request, next(get_db()))
    if user:
        if user.role == "super_admin":
            return RedirectResponse(url="/admin/dashboard", status_code=302)
        elif user.role == "company_admin":
            return RedirectResponse(url="/receipt/dashboard", status_code=302)
        else:
            return RedirectResponse(url="/review/dashboard", status_code=302)
    return RedirectResponse(url="/login", status_code=302)

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/api/login")
async def login(
    response: Response,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.username == username).first()

    if not user or user.password_hash != hash_password(password):
        return HTMLResponse(content="""
            <script>
                alert('아이디 또는 비밀번호가 올바르지 않습니다');
                window.location.href = '/login';
            </script>
        """)

    if not user.is_active:
        return HTMLResponse(content="""
            <script>
                alert('비활성화된 계정입니다');
                window.location.href = '/login';
            </script>
        """)

    session_id = create_session(user.id)

    if user.role == "super_admin":
        redirect_url = "/admin/dashboard"
    elif user.role == "company_admin":
        redirect_url = "/receipt/dashboard"
    else:
        redirect_url = "/review/dashboard"

    response = RedirectResponse(url=redirect_url, status_code=302)
    response.set_cookie(key="session_id", value=session_id, httponly=True)
    return response

@app.get("/logout")
async def logout(response: Response, request: Request):
    session_id = request.cookies.get("session_id")
    if session_id and session_id in sessions:
        del sessions[session_id]

    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie(key="session_id")
    return response

# 페이지 라우트
@app.get("/receipt/order", response_class=HTMLResponse)
async def receipt_order_page(request: Request, user = Depends(require_login)):
    return templates.TemplateResponse("receipt_order_form.html", {"request": request, "user": user})

@app.get("/receipt/dashboard", response_class=HTMLResponse)
async def receipt_dashboard(request: Request, user = Depends(require_login), db: Session = Depends(get_db)):
    # 고객사 주문 목록 가져오기 (리뷰 정보 포함)
    orders = db.query(ReceiptWorkOrder).options(
        joinedload(ReceiptWorkOrder.reviews)
    ).filter(ReceiptWorkOrder.client_id == user.id).all()

    # 헬퍼 함수: 성공적으로 추출된 리뷰인지 확인
    def is_valid_review(review):
        """리뷰가 정상적으로 추출되었는지 확인 (content와 receipt_date_str이 모두 유효한 경우)"""
        if not review.content or not review.receipt_date_str:
            return False

        # 에러 메시지가 포함된 경우 필터링
        error_keywords = ['내용 추출 대기중', '찾을 수 없습니다', '내용 없음', '에러', 'error']
        content_lower = review.content.lower()
        receipt_lower = review.receipt_date_str.lower()

        for keyword in error_keywords:
            if keyword in content_lower or keyword in receipt_lower:
                return False

        return True

    # 고객사 계정: 각 주문에 대해 성공적으로 추출된 리뷰만 필터링
    for order in orders:
        # 원본 리뷰는 유지하되, 유효한 리뷰만 필터링
        valid_reviews = [r for r in order.reviews if is_valid_review(r)]
        # 고객사에게 보여줄 리뷰 목록으로 덮어쓰기
        order.reviews = valid_reviews
        # completed_count도 유효한 리뷰 개수로 업데이트
        order.completed_count = len(valid_reviews)

    # 통계 계산
    stats = {
        'pending': len([o for o in orders if o.status == 'pending']),
        'approved': len([o for o in orders if o.status == 'approved']),
        'completed': len([o for o in orders if o.status == 'completed']),
        'total': len(orders)
    }

    # 추가 통계 계산
    total_orders = len(orders)
    completed_orders_list = [o for o in orders if o.status == 'completed']
    in_progress_orders = len([o for o in orders if o.status == 'in_progress'])
    pending_orders = len([o for o in orders if o.status == 'pending'])
    completed_orders = len(completed_orders_list)
    recent_orders = sorted(orders, key=lambda x: x.created_at, reverse=True)

    # 사용 이력 계산
    from datetime import datetime, timedelta, date
    now = datetime.now()
    current_month_start = date(now.year, now.month, 1)
    if now.month == 1:
        last_month_start = date(now.year - 1, 12, 1)
    else:
        last_month_start = date(now.year, now.month - 1, 1)

    # 이번 달 사용 금액
    current_month_total = sum(
        o.total_price for o in orders
        if o.status == 'completed' and o.start_date and o.start_date >= current_month_start
    )

    # 지난 달 사용 금액
    last_month_total = sum(
        o.total_price for o in orders
        if o.status == 'completed' and o.start_date and
        o.start_date >= last_month_start and o.start_date < current_month_start
    )

    # 전체 누적 사용 금액
    total_usage = sum(o.total_price for o in orders if o.status == 'completed')

    return templates.TemplateResponse("client_dashboard_simple.html", {
        "request": request,
        "user": user,
        "current_user": user,  # 템플릿에서 current_user 사용을 위해 추가
        "orders": orders,
        "stats": stats,
        "total_orders": total_orders,
        "completed_orders": completed_orders,
        "in_progress_orders": in_progress_orders,
        "pending_orders": pending_orders,
        "recent_orders": recent_orders,
        "completed_orders": completed_orders_list,
        "current_month_total": current_month_total,
        "last_month_total": last_month_total,
        "total_usage": total_usage
    })

@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request, user = Depends(require_super_admin), db: Session = Depends(get_db)):
    # 모든 주문 가져오기
    orders = db.query(ReceiptWorkOrder).order_by(ReceiptWorkOrder.created_at.desc()).all()

    # 모든 리뷰 가져오기
    all_reviews = db.query(Review).order_by(Review.created_at.desc()).all()

    # 통계 계산
    stats = {
        'pending': len([o for o in orders if o.status == 'pending']),
        'approved': len([o for o in orders if o.status == 'approved']),
        'completed': len([o for o in orders if o.status == 'completed']),
        'total': len(orders)
    }

    return templates.TemplateResponse("admin_dashboard_pro.html", {
        "request": request,
        "user": user,
        "orders": orders,
        "all_reviews": all_reviews,
        "stats": stats
    })

@app.get("/review/dashboard", response_class=HTMLResponse)
async def review_dashboard(request: Request, user = Depends(require_login)):
    return templates.TemplateResponse("review_dashboard.html", {"request": request, "user": user})

# 고객사 관리 페이지
@app.get("/admin/clients", response_class=HTMLResponse)
async def admin_clients_page(request: Request, user = Depends(require_super_admin)):
    return templates.TemplateResponse("admin_clients.html", {"request": request, "user": user})

# 고객사 목록 조회 API
@app.get("/api/admin/clients")
async def get_clients(user = Depends(require_super_admin), db: Session = Depends(get_db)):
    clients = db.query(User).filter(User.role == "company_admin").all()

    result = []
    for client in clients:
        # 해당 클라이언트의 주문 수 계산
        order_count = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.client_id == client.id).count()

        result.append({
            "id": client.id,
            "username": client.username,
            "company_name": client.company.name if client.company else "",
            "full_name": client.full_name,
            "phone": client.phone,
            "email": client.email,
            "unit_price": client.company.unit_price if client.company else 3000,
            "created_at": client.created_at.isoformat() if client.created_at else "",
            "order_count": order_count
        })

    return result

# 고객사 등록 API
@app.post("/api/admin/clients")
async def create_client(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    try:
        data = await request.json()

        # 중복 체크
        existing = db.query(User).filter(User.username == data['username']).first()
        if existing:
            return {"success": False, "message": "이미 존재하는 아이디입니다."}

        # 회사 생성
        company_name = data.get('company_name', '')
        company = Company(
            name=company_name,
            display_name=company_name,  # display_name도 같은 값으로 설정
            unit_price=float(data['unit_price'])  # 단가는 필수 입력
        )
        db.add(company)
        db.flush()

        # 사용자 생성
        new_user = User(
            username=data['username'],
            password_hash=hashlib.sha256(data['password'].encode()).hexdigest(),
            full_name=data.get('contact_name', ''),
            phone=data.get('phone', ''),
            email=data.get('email', ''),
            role="company_admin",
            company_id=company.id
        )
        db.add(new_user)
        db.commit()

        return {"success": True, "message": "고객사가 등록되었습니다."}
    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

# 고객사 삭제 API
@app.delete("/api/admin/clients/{client_id}")
async def delete_client(
    client_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    try:
        client = db.query(User).filter(User.id == client_id).first()
        if not client:
            return {"success": False, "message": "고객사를 찾을 수 없습니다."}

        # 관련 데이터 삭제 (주문, 리뷰 등)
        orders = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.client_id == client_id).all()
        for order in orders:
            # 리뷰 삭제
            db.query(Review).filter(Review.order_id == order.id).delete()
            # 주문 삭제
            db.delete(order)

        # 회사 삭제
        if client.company_id:
            company = db.query(Company).filter(Company.id == client.company_id).first()
            if company:
                db.delete(company)

        # 사용자 삭제
        db.delete(client)
        db.commit()

        return {"success": True, "message": "고객사가 삭제되었습니다."}
    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

# 고객사 수정 API
@app.put("/api/admin/clients/{client_id}")
async def update_client(
    client_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    try:
        data = await request.json()

        # 클라이언트 찾기
        client = db.query(User).filter(User.id == client_id).first()
        if not client:
            return {"success": False, "message": "고객사를 찾을 수 없습니다."}

        # 회사 정보 찾기
        company = db.query(Company).filter(Company.id == client.company_id).first()
        if not company:
            return {"success": False, "message": "회사 정보를 찾을 수 없습니다."}

        # 아이디 중복 체크 (현재 클라이언트 제외)
        if 'username' in data and data['username'] != client.username:
            existing = db.query(User).filter(
                User.username == data['username'],
                User.id != client_id
            ).first()
            if existing:
                return {"success": False, "message": "이미 존재하는 아이디입니다."}

        # 회사 정보 업데이트
        if 'company_name' in data:
            company.name = data['company_name']
            company.display_name = data['company_name']
        if 'unit_price' in data:
            company.unit_price = float(data['unit_price'])

        # 사용자 정보 업데이트
        if 'username' in data:
            client.username = data['username']
        if 'password' in data and data['password']:  # 비밀번호가 있을 때만 업데이트
            client.password_hash = hashlib.sha256(data['password'].encode()).hexdigest()
        if 'contact_name' in data:
            client.full_name = data['contact_name']
        if 'phone' in data:
            client.phone = data['phone']
        if 'email' in data:
            client.email = data['email']

        db.commit()
        return {"success": True, "message": "고객사 정보가 수정되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

# 영수증 작업 요청 API
@app.post("/api/receipt/order")
async def create_receipt_order(
    request: Request,
    business_name: str = Form(...),
    representative_name: str = Form(...),
    business_number: str = Form(...),
    business_type: str = Form(...),
    place_number: str = Form(...),
    place_link: str = Form(...),
    business_address: str = Form(...),
    working_days: int = Form(...),
    daily_count: int = Form(...),
    receipt_date: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    total_count: int = Form(...),
    guidelines: str = Form(None),
    attachment_images: list[UploadFile] = File(...),
    review_excel: UploadFile = File(None),
    review_photos: UploadFile = File(None),
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    try:
        now = datetime.now()
        order_no = f"RC{now.strftime('%Y%m%d%H%M%S')}"

        # 이미지 파일 저장
        upload_dir = "uploads/orders"
        os.makedirs(upload_dir, exist_ok=True)

        saved_image_paths = []
        for idx, file in enumerate(attachment_images):
            # 파일 확장자 확인
            file_ext = file.filename.split('.')[-1].lower()
            if file_ext not in ['jpg', 'jpeg', 'png']:
                return JSONResponse({
                    "success": False,
                    "message": f"지원하지 않는 파일 형식입니다: {file.filename}"
                }, status_code=400)

            # 파일명 생성
            file_name = f"{order_no}_{idx+1}.{file_ext}"
            file_path = os.path.join(upload_dir, file_name)

            # 파일 저장
            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)

            saved_image_paths.append(file_path)

        # 필수 필드 검증
        if not place_number.strip():
            return JSONResponse({
                "success": False,
                "message": "플레이스 등록한 번호는 필수 입력 항목입니다."
            }, status_code=400)

        if not place_link.strip():
            return JSONResponse({
                "success": False,
                "message": "플레이스 링크는 필수 입력 항목입니다."
            }, status_code=400)

        # 날짜 파싱
        receipt_date_obj = datetime.strptime(receipt_date, '%Y-%m-%d').date()
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 리뷰 자료 파일 저장 (선택사항)
        review_excel_path = None
        review_photos_path = None

        if review_excel and review_excel.filename:
            review_dir = "uploads/review_assets"
            os.makedirs(review_dir, exist_ok=True)
            excel_filename = f"{order_no}_review.xlsx"
            excel_path = os.path.join(review_dir, excel_filename)
            with open(excel_path, "wb") as f:
                content = await review_excel.read()
                f.write(content)
            review_excel_path = excel_path

        if review_photos and review_photos.filename:
            review_dir = "uploads/review_assets"
            os.makedirs(review_dir, exist_ok=True)
            zip_filename = f"{order_no}_photos.zip"
            zip_path = os.path.join(review_dir, zip_filename)
            with open(zip_path, "wb") as f:
                content = await review_photos.read()
                f.write(content)
            review_photos_path = zip_path

        # 고객사별 단가 가져오기
        company = db.query(Company).filter(Company.id == user.company_id).first()
        unit_price = company.unit_price if company else 3000
        total_price = total_count * unit_price

        new_order = ReceiptWorkOrder(
            order_no=order_no,
            company_id=user.company_id if user.company_id else 1,
            client_id=user.id,
            business_name=business_name,
            representative_name=representative_name,
            business_number=business_number,
            business_type=business_type,
            place_number=place_number,
            place_link=place_link,
            business_address=business_address,
            receipt_date=receipt_date_obj,
            start_date=start_date_obj,
            end_date=end_date_obj,
            daily_count=daily_count,
            working_days=working_days,
            total_count=total_count,
            unit_price=unit_price,
            total_price=total_price,
            guidelines=guidelines if guidelines else '',
            attachment_images=json.dumps(saved_image_paths),
            review_excel_path=review_excel_path,
            review_photos_path=review_photos_path
        )

        db.add(new_order)
        db.commit()

        return JSONResponse({
            "success": True,
            "order_no": order_no,
            "message": "작업 요청이 접수되었습니다"
        })

    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"작업 처리 중 오류가 발생했습니다: {str(e)}"
        }, status_code=500)

# 고객사 대량 업체 등록
@app.post("/api/client/bulk-register")
async def client_bulk_register(
    file: UploadFile = File(...),
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """고객사가 여러 업체를 한번에 등록"""
    try:
        content = await file.read()
        df = read_file_to_dataframe(content, file.filename)

        log = []
        count = 0

        for _, row in df.iterrows():
            try:
                order_no = f"RC{datetime.now().strftime('%Y%m%d%H%M%S')}_{count}"
                receipt_date = date.today()

                if receipt_date.weekday() == 4:
                    start_date = receipt_date + timedelta(days=3)
                else:
                    start_date = receipt_date + timedelta(days=1)

                working_days = int(row.get('작업일수', 1))
                daily_count = int(row.get('일일작업량', 10))
                end_date = start_date + timedelta(days=working_days - 1)
                total_count = daily_count * working_days

                company = db.query(Company).filter(Company.id == user.company_id).first()
                unit_price = company.unit_price if company else 3000
                total_price = total_count * unit_price

                new_order = ReceiptWorkOrder(
                    order_no=order_no,
                    company_id=user.company_id,
                    client_id=user.id,
                    business_name=str(row['업체명']),
                    representative_name=str(row['대표자명']),
                    business_number=str(row['사업자번호']),
                    place_number=str(row.get('플레이스번호', '')),
                    business_address=str(row['주소']),
                    receipt_date=receipt_date,
                    start_date=start_date,
                    end_date=end_date,
                    daily_count=daily_count,
                    working_days=working_days,
                    total_count=total_count,
                    unit_price=unit_price,
                    total_price=total_price,
                    guidelines=str(row.get('가이드라인', ''))
                )

                db.add(new_order)
                count += 1
                log.append(f"[OK] {row['업체명']} 등록 완료")

            except Exception as e:
                log.append(f"[ERROR] {row.get('업체명', 'Unknown')} 등록 실패: {str(e)}")

        db.commit()
        return {"success": True, "count": count, "log": log}

    except Exception as e:
        return {"success": False, "message": str(e)}

# 고객사 작업 주문 조회
@app.get("/api/company/orders")
async def get_company_orders(
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    if user.role == "super_admin":
        orders = db.query(ReceiptWorkOrder).order_by(ReceiptWorkOrder.created_at.desc()).all()
    else:
        orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.client_id == user.id
        ).order_by(ReceiptWorkOrder.created_at.desc()).all()

    result = []
    for order in orders:
        # 리뷰 수 카운트 - 추출 실패한 리뷰는 제외
        review_count = db.query(Review).filter(
            Review.order_id == order.id,
            ~Review.content.like("추출 실패%")
        ).count()

        # 실제 추출 완료된 리뷰 수 계산
        extracted_count = db.query(Review).filter(
            Review.order_id == order.id,
            Review.content != "내용 추출 대기중",
            Review.content != None,
            Review.content != "",
            ~Review.content.like("추출 실패%")
        ).count()

        # completed_count와 status 실시간 업데이트
        if extracted_count != order.completed_count:
            order.completed_count = extracted_count
            db.commit()

        if extracted_count >= order.total_count and order.status == 'approved':
            order.status = 'completed'
            order.completed_at = datetime.now()
            db.commit()

        # 완료된 주문은 항상 100%로 표시
        if order.status == 'completed':
            progress_rate = 100.0
        else:
            progress_rate = (order.completed_count / order.total_count * 100) if order.total_count > 0 else 0

        result.append({
            "id": order.id,
            "order_no": order.order_no,
            "business_name": order.business_name,
            "representative_name": order.representative_name,
            "place_number": order.place_number,
            "place_link": order.place_link,
            "start_date": order.start_date.isoformat(),
            "end_date": order.end_date.isoformat(),
            "daily_count": order.daily_count,
            "working_days": order.working_days,
            "total_count": order.total_count,
            "completed_count": order.completed_count,
            "review_count": review_count,
            "progress_rate": round(progress_rate, 1),
            "status": order.status,
            "unit_price": int(order.unit_price) if order.unit_price else 0,
            "total_price": int(order.total_price) if order.total_price else 0,
            "rejection_reason": order.rejection_reason if order.status == "rejected" else None,
            "created_at": order.created_at.isoformat()
        })

    return {"success": True, "data": result}

# 관리자 - 모든 작업 주문 조회
@app.get("/api/admin/orders")
async def get_all_orders(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    orders = db.query(ReceiptWorkOrder).order_by(ReceiptWorkOrder.created_at.desc()).all()

    result = []
    for order in orders:
        # 리뷰 수 카운트 - 추출 실패한 리뷰는 제외
        review_count = db.query(Review).filter(
            Review.order_id == order.id,
            ~Review.content.like("추출 실패%")
        ).count()

        # 실제 추출 완료된 리뷰 수 계산
        extracted_count = db.query(Review).filter(
            Review.order_id == order.id,
            Review.content != "내용 추출 대기중",
            Review.content != None,
            Review.content != "",
            ~Review.content.like("추출 실패%")
        ).count()

        # completed_count와 status 실시간 업데이트
        if extracted_count != order.completed_count:
            order.completed_count = extracted_count
            db.commit()

        if extracted_count >= order.total_count and order.status == 'approved':
            order.status = 'completed'
            order.completed_at = datetime.now()
            db.commit()

        # 완료된 주문은 항상 100%로 표시
        if order.status == 'completed':
            progress_rate = 100.0
        else:
            progress_rate = (order.completed_count / order.total_count * 100) if order.total_count > 0 else 0

        # 연장 요청 확인 (비활성화)
        extension = None  # 연장 기능 제거
        # extension = db.query(ExtensionRequest).filter(
        #     ExtensionRequest.order_id == order.id,
        #     ExtensionRequest.status == 'pending'
        # ).first()

        result.append({
            "id": order.id,
            "order_no": order.order_no,
            "company_name": order.company.display_name if order.company else "알 수 없음",
            "client_company": order.company.display_name if order.company else "알 수 없음",
            "client_username": order.client.username if order.client else "알 수 없음",
            "business_name": order.business_name,
            "representative_name": order.representative_name,
            "place_number": order.place_number,
            "place_link": order.place_link,
            "start_date": order.start_date.isoformat() if order.start_date else "",
            "end_date": order.end_date.isoformat() if order.end_date else "",
            "daily_count": order.daily_count,
            "working_days": order.working_days,
            "total_count": order.total_count,
            "completed_count": order.completed_count,
            "review_count": review_count,
            "unit_price": int(order.unit_price) if order.unit_price else 0,
            "progress_rate": round(progress_rate, 1),
            "status": order.status,
            "total_price": int(order.total_price) if order.total_price else 0,
            "has_extension": False,  # 연장 기능 제거
            "extension_id": None,
            "extension_days": None,
            "extension_daily": None,
            "created_at": order.created_at.isoformat()
        })

    return {"success": True, "data": result}

# 작업 승인
@app.put("/api/admin/orders/{order_id}/approve")
async def approve_order(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

    if not order:
        return {"success": False, "message": "주문을 찾을 수 없습니다"}

    order.status = "approved"
    order.approved_at = datetime.now()
    order.approved_by = user.id
    db.commit()

    return {"success": True, "message": "주문이 승인되었습니다"}

# 작업 주문 수정 API
@app.put("/api/admin/orders/{order_id}/edit")
async def edit_order(
    order_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """관리자가 작업 주문 정보를 수정"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

        if not order:
            return JSONResponse({
                "success": False,
                "message": "주문을 찾을 수 없습니다"
            }, status_code=404)

        data = await request.json()

        # 수정 가능한 필드들
        if 'start_date' in data:
            order.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()

        if 'end_date' in data:
            order.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

        if 'daily_count' in data:
            order.daily_count = int(data['daily_count'])
            # total_count 재계산
            if order.working_days:
                order.total_count = order.daily_count * order.working_days
                order.total_price = order.total_count * order.unit_price

        if 'working_days' in data:
            order.working_days = int(data['working_days'])
            # end_date와 total_count 재계산
            if order.start_date:
                order.end_date = order.start_date + timedelta(days=order.working_days - 1)
            order.total_count = order.daily_count * order.working_days
            order.total_price = order.total_count * order.unit_price

        if 'business_name' in data:
            order.business_name = data['business_name']

        if 'representative_name' in data:
            order.representative_name = data['representative_name']

        if 'place_number' in data:
            order.place_number = data['place_number']

        if 'place_link' in data:
            order.place_link = data['place_link']

        db.commit()

        return JSONResponse({
            "success": True,
            "message": "주문 정보가 수정되었습니다"
        })

    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"수정 중 오류가 발생했습니다: {str(e)}"
        }, status_code=500)

# 작업 거절
@app.put("/api/admin/orders/{order_id}/reject")
async def reject_order(
    order_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    data = await request.json()
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

    if not order:
        return {"success": False, "message": "주문을 찾을 수 없습니다"}

    order.status = "rejected"
    order.rejection_reason = data.get("reason", "관리자가 거절함")
    order.approved_at = datetime.now()
    order.approved_by = user.id
    db.commit()

    return {"success": True, "message": "주문이 거절되었습니다"}

# 업체별 리뷰 URL 목록 가져오기
@app.get("/api/admin/orders/{order_id}/reviews")
async def get_order_reviews(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """특정 주문의 리뷰 URL 목록 가져오기"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()
        if not order:
            return JSONResponse({
                "success": False,
                "message": "주문을 찾을 수 없습니다"
            }, status_code=404)

        reviews = db.query(Review).filter(Review.order_id == order_id).all()

        return JSONResponse({
            "success": True,
            "reviews": [{
                "id": review.id,
                "review_url": review.review_url,
                "content": review.content,
                "extracted_at": review.extracted_at.isoformat() if review.extracted_at else None
            } for review in reviews]
        })
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"오류: {str(e)}"
        }, status_code=500)

@app.delete("/api/admin/orders/{order_id}/reviews")
async def delete_order_reviews_only(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """업체의 리뷰만 전체 삭제 (주문은 유지)"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

        if not order:
            return JSONResponse({"success": False, "message": "주문을 찾을 수 없습니다."}, status_code=404)

        # 리뷰 개수 확인
        review_count = db.query(Review).filter(Review.order_id == order_id).count()

        if review_count == 0:
            return {"success": True, "message": "삭제할 리뷰가 없습니다."}

        # 모든 리뷰 삭제
        db.query(Review).filter(Review.order_id == order_id).delete()

        # 주문의 completed_count 초기화
        order.completed_count = 0

        # 완료 상태였다면 진행중으로 변경
        if order.status == 'completed':
            order.status = 'approved'
            order.completed_at = None

        db.commit()

        return {"success": True, "message": f"{review_count}개의 리뷰가 삭제되었습니다. (주문은 유지됨)"}

    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# 업체별 리뷰 URL 대량 추가
@app.post("/api/admin/orders/{order_id}/reviews/bulk")
async def add_order_reviews_bulk(
    order_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """특정 주문에 여러 리뷰 URL 한번에 추가"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()
        if not order:
            return JSONResponse({
                "success": False,
                "message": "주문을 찾을 수 없습니다"
            }, status_code=404)

        data = await request.json()
        review_urls = data.get("review_urls", [])

        if not review_urls:
            return JSONResponse({
                "success": False,
                "message": "리뷰 URL이 없습니다"
            }, status_code=400)

        added_count = 0
        for url in review_urls:
            url = url.strip()
            if not url:
                continue

            # 중복 체크
            existing = db.query(Review).filter(
                Review.order_id == order_id,
                Review.review_url == url
            ).first()

            if existing:
                continue

            # 새 리뷰 추가
            review = Review(
                order_id=order_id,
                review_url=url,
                content="내용 추출 대기중",
                rating=5,
                review_date=datetime.now().date()
            )
            db.add(review)
            added_count += 1

        db.commit()

        return JSONResponse({
            "success": True,
            "message": f"{added_count}개의 리뷰 URL이 추가되었습니다",
            "count": added_count
        })

    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"추가 중 오류: {str(e)}"
        }, status_code=500)

# 업체별 리뷰 추출
@app.post("/api/admin/orders/{order_id}/extract-reviews")
async def extract_order_reviews(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """특정 주문의 모든 리뷰 추출 (기존 시스템 로직 그대로 사용)"""
    try:
        from real_review_extractor import get_extractor
        extractor = get_extractor()

        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()
        if not order:
            return JSONResponse({
                "success": False,
                "message": "주문을 찾을 수 없습니다"
            }, status_code=404)

        reviews = db.query(Review).filter(Review.order_id == order_id).all()

        if not reviews:
            return JSONResponse({
                "success": False,
                "message": "추출할 리뷰가 없습니다"
            }, status_code=400)

        success_count = 0
        fail_count = 0
        shop_name = order.business_name

        for review in reviews:
            try:
                if not review.review_url:
                    fail_count += 1
                    continue

                # 이미 추출된 경우 건너뛰기
                if review.extracted_at and review.content:
                    continue

                # 이전 내용 저장 (카운팅 판단용)
                old_content = review.content

                # 실제 리뷰 내용 추출
                review_text, receipt_date, metadata = extractor.extract_review(
                    review.review_url,
                    shop_name
                )

                if review_text and "오류" not in review_text and "찾을 수 없습니다" not in review_text:
                    review.content = review_text
                    review.receipt_date_str = receipt_date
                    review.extracted_at = datetime.now()

                    # 처음 추출 성공한 경우에만 카운팅 (기존 로직 그대로)
                    if old_content == "내용 추출 대기중":
                        # 영수증 날짜가 있는 경우만 카운팅 (필수 조건)
                        if receipt_date:
                            # 중복 체크 - 같은 업체명, 같은 URL, 같은 내용이 이미 추출된 경우
                            existing = db.query(Review).join(ReceiptWorkOrder).filter(
                                Review.id != review.id,
                                ReceiptWorkOrder.business_name == order.business_name,
                                Review.review_url == review.review_url,
                                Review.content == review_text,
                                Review.content != "내용 추출 대기중",
                                Review.content != None,
                                Review.content != ""
                            ).first()

                            if existing:
                                # 중복이면 현재 리뷰 삭제
                                print(f"[중복 발견] 리뷰 {review.id} - 동일 업체명/URL/내용 (삭제)")
                                db.delete(review)
                                fail_count += 1
                            else:
                                # 중복이 아닌 경우만 카운팅
                                if order.completed_count < order.total_count:
                                    order.completed_count += 1
                                    print(f"리뷰 {review.id} 추출 완료 (영수증 날짜: {receipt_date}) - 카운팅: {order.completed_count}/{order.total_count}")
                                else:
                                    print(f"[초과 리뷰] 리뷰 {review.id} - completed_count가 이미 total_count 초과")
                                success_count += 1
                        else:
                            print(f"[영수증 날짜 없음] 리뷰 {review.id} - 영수증 날짜를 찾을 수 없어 카운팅 제외")
                            fail_count += 1

                        # 주문 완료 체크
                        if order.completed_count >= order.total_count and order.status != 'completed':
                            order.status = 'completed'
                            order.completed_at = datetime.now()
                            print(f"[완료] 주문 {order.id} ({order.business_name}) - 상태 completed로 변경")

                    db.commit()
                else:
                    print(f"리뷰 추출 실패 ({review.id}): {review_text}")
                    fail_count += 1

            except Exception as e:
                print(f"리뷰 추출 실패 ({review.id}): {e}")
                fail_count += 1
                db.rollback()

        return JSONResponse({
            "success": True,
            "message": f"추출 완료: 성공 {success_count}건, 실패 {fail_count}건",
            "success_count": success_count,
            "fail_count": fail_count
        })

    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"추출 중 오류: {str(e)}"
        }, status_code=500)

# 엑셀 다운로드 (미승인 작업만)
@app.get("/api/admin/export/excel")
async def export_excel(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """미승인 작업만 엑셀로 내보내기"""
    try:
        orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.status == "pending"
        ).all()

        data = []
        for order in orders:
            data.append({
                '주문번호': order.order_no,
                '고객사': order.company.display_name if order.company else "",
                '업체명': order.business_name,
                '대표자': order.representative_name,
                '사업자번호': order.business_number,
                '주소': order.business_address,
                '시작일': order.start_date.strftime('%Y-%m-%d'),
                '종료일': order.end_date.strftime('%Y-%m-%d'),
                '일일작업량': order.daily_count,
                '총작업량': order.total_count,
                '단가': order.unit_price,
                '총금액': order.total_price,
                '등록일': order.created_at.strftime('%Y-%m-%d %H:%M'),
                '요청자': order.client.full_name if order.client else ""
            })

        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='미승인작업')

        output.seek(0)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''pending_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.get("/api/admin/export/report")
async def export_report(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """완료된 작업 리포트 엑셀로 내보내기"""
    try:
        orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.status == "approved"
        ).all()

        data = []
        for order in orders:
            # 리뷰 수 계산
            review_count = db.query(Review).filter(Review.order_id == order.id).count()

            data.append({
                '주문번호': order.order_no,
                '고객사': order.company.display_name if order.company else "",
                '업체명': order.business_name,
                '대표자': order.representative_name,
                '사업자번호': order.business_number,
                '주소': order.business_address,
                '시작일': order.start_date.strftime('%Y-%m-%d'),
                '종료일': order.end_date.strftime('%Y-%m-%d'),
                '일일작업량': order.daily_count,
                '총작업량': order.total_count,
                '리뷰수': review_count,
                '단가': order.unit_price,
                '총금액': order.total_price,
                '상태': '완료',
                '등록일': order.created_at.strftime('%Y-%m-%d %H:%M'),
                '승인일': order.updated_at.strftime('%Y-%m-%d %H:%M') if order.updated_at else "",
                '요청자': order.client.full_name if order.client else ""
            })

        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='완료작업')

        output.seek(0)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# 고객사 관리 API
@app.get("/api/admin/clients")
async def get_clients(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    try:
        clients = db.query(User).filter(User.role == "company_admin").all()
        result = []
        for client in clients:
            company = db.query(Company).filter(Company.id == client.company_id).first()
            manager = None
            if company and company.manager_id:
                manager = db.query(Manager).filter(Manager.id == company.manager_id).first()
            # 해당 클라이언트의 주문 수 계산
            order_count = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.client_id == client.id).count()

            result.append({
                "id": client.id,
                "username": client.username,
                "company_name": company.display_name if company else "Unknown",
                "full_name": client.full_name,
                "phone": client.phone,
                "email": client.email,
                "created_at": client.created_at.strftime('%Y-%m-%d'),
                "order_count": order_count,
                "unit_price": company.unit_price if company else None,
                "manager_name": manager.name if manager else None,
                "manager_id": company.manager_id if company else None
            })
        return result
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/admin/clients/add")
async def add_client(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    try:
        data = await request.json()

        # 중복 확인
        existing_user = db.query(User).filter(User.username == data['username']).first()
        if existing_user:
            return {"success": False, "message": "이미 존재하는 아이디입니다"}

        # 회사 생성
        new_company = Company(
            name=data['name'].replace(" ", "_").lower(),
            display_name=data['display_name'],
            unit_price=float(data['unit_price']),
            manager_id=data.get('manager_id')  # 담당자 ID 추가
        )
        db.add(new_company)
        db.flush()

        # 고객사 관리자 계정 생성
        new_user = User(
            username=data['username'],
            password_hash=hash_password(data['password']),
            full_name=data['display_name'],
            role="company_admin",
            company_id=new_company.id
        )
        db.add(new_user)
        db.commit()

        return {"success": True, "message": "고객사가 추가되었습니다"}
    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"추가 실패: {str(e)}"}

@app.post("/api/admin/clients/register")
async def register_client(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    data = await request.json()

    # 회사 생성
    new_company = Company(
        name=data['company_name'].replace(" ", "_").lower(),
        display_name=data['company_name'],
        unit_price=float(data.get('unit_price', 3000))
    )
    db.add(new_company)
    db.flush()

    # 관리자 계정 생성
    new_user = User(
        username=data['username'],
        password_hash=hash_password(data['password']),
        full_name=data['contact_name'],
        role="company_admin",
        company_id=new_company.id
    )
    db.add(new_user)
    db.commit()

    return {"success": True, "message": "고객사가 등록되었습니다"}

# 리뷰어 관리 API
@app.get("/api/admin/reviewers")
async def get_reviewers(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    reviewers = db.query(User).filter(User.role == "reviewer").all()

    result = []
    for reviewer in reviewers:
        review_count = db.query(Review).filter(Review.reviewer_id == reviewer.id).count()
        result.append({
            "id": reviewer.id,
            "full_name": reviewer.full_name,
            "username": reviewer.username,
            "company_name": reviewer.company.display_name if reviewer.company else "미지정",
            "total_reviews": review_count,
            "email": reviewer.email,
            "is_active": reviewer.is_active
        })

    return {"success": True, "data": result}

@app.post("/api/admin/reviewers/register")
async def register_reviewer(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    data = await request.json()

    new_reviewer = User(
        username=data['username'],
        password_hash=hash_password(data['password']),
        full_name=data['full_name'],
        email=data.get('email', ''),
        role="reviewer",
        company_id=int(data['company_id']) if data.get('company_id') else None
    )
    db.add(new_reviewer)
    db.commit()

    return {"success": True, "message": "리뷰어가 등록되었습니다"}

# 인센티브 대시보드 API
@app.get("/api/admin/incentive-report")
async def get_incentive_report(
    period: str = "7",  # 1, 7, 30 or custom
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """담당자별 매출 및 인센티브 리포트"""

    # 기간 설정
    if period == "custom" and start_date and end_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    else:
        end = date.today()
        days = int(period)
        start = end - timedelta(days=days)

    # 담당자별 매출 집계
    results = []

    managers = db.query(User).filter(User.role == "company_admin").all()

    for manager in managers:
        # 해당 기간 동안의 승인된 주문들
        orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.client_id == manager.id,
            ReceiptWorkOrder.status.in_(["approved", "in_progress", "completed"]),
            ReceiptWorkOrder.created_at >= start,
            ReceiptWorkOrder.created_at <= datetime.combine(end, datetime.max.time())
        ).all()

        total_revenue = sum(order.total_price for order in orders)
        total_orders = len(orders)
        total_count = sum(order.total_count for order in orders)
        completed_count = sum(order.completed_count for order in orders)

        # 인센티브 계산 (완료된 작업 기준)
        incentive_rate = manager.commission_rate or 0.1
        incentive = total_revenue * incentive_rate

        results.append({
            "manager_id": manager.id,
            "manager_name": manager.full_name,
            "company_name": manager.company.display_name if manager.company else "미지정",
            "total_orders": total_orders,
            "total_count": total_count,
            "completed_count": completed_count,
            "total_revenue": total_revenue,
            "incentive_rate": incentive_rate * 100,
            "incentive": incentive,
            "period": f"{start} ~ {end}"
        })

    # 총계
    summary = {
        "total_revenue": sum(r["total_revenue"] for r in results),
        "total_incentive": sum(r["incentive"] for r in results),
        "total_orders": sum(r["total_orders"] for r in results),
        "managers": results
    }

    return {"success": True, "data": summary}

# 대량 업로드 - 고객사
@app.post("/api/admin/upload/clients")
async def bulk_upload_clients(
    file: UploadFile = File(...),
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """고객사 대량 업로드"""
    try:
        content = await file.read()
        df = read_file_to_dataframe(content, file.filename)

        log = []
        count = 0

        for _, row in df.iterrows():
            try:
                # 회사 생성
                company = Company(
                    name=row['회사명'].replace(" ", "_").lower(),
                    display_name=row['회사명'],
                    unit_price=float(row.get('단가', 3000))
                )
                db.add(company)
                db.flush()

                # 관리자 계정 생성
                user_obj = User(
                    username=row['아이디'],
                    password_hash=hash_password(row['비밀번호']),
                    full_name=row['담당자'],
                    role="company_admin",
                    company_id=company.id
                )
                db.add(user_obj)
                count += 1
                log.append(f"[OK] {row['회사명']} 등록 완료")

            except Exception as e:
                log.append(f"[ERROR] {row.get('회사명', 'Unknown')} 등록 실패: {str(e)}")

        db.commit()
        return {"success": True, "count": count, "log": log}

    except Exception as e:
        return {"success": False, "message": str(e)}

# 대량 업로드 - 리뷰
@app.post("/api/admin/upload/reviews")
async def bulk_upload_reviews(
    file: UploadFile = File(...),
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """리뷰 대량 업로드"""
    try:
        content = await file.read()
        df = read_file_to_dataframe(content, file.filename)

        log = []
        count = 0

        # 엑셀 파일 컬럼 확인
        log.append(f"[INFO] 업로드된 파일 컬럼: {list(df.columns)}")
        log.append(f"[INFO] 총 {len(df)}개의 행 발견")

        for idx, row in df.iterrows():
            try:
                # 업체명 확인
                business_name = row.get('업체명', '')
                if not business_name:
                    log.append(f"[SKIP] 행 {idx+1}: 업체명이 비어있음")
                    continue

                # 모든 주문 상태 확인 (디버깅용)
                all_orders = db.query(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.business_name == business_name
                ).all()

                if not all_orders:
                    log.append(f"[ERROR] '{business_name}'에 대한 주문이 전혀 없음")
                else:
                    status_list = [f"{o.status}({o.completed_count}/{o.total_count})" for o in all_orders]
                    log.append(f"[INFO] '{business_name}' 주문 현황: {', '.join(status_list)}")

                # 업체명으로 미완료된 승인 주문 찾기 (가장 오래된 미완료 주문부터)
                order = db.query(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.business_name == business_name,
                    ReceiptWorkOrder.status == 'approved',
                    ReceiptWorkOrder.completed_count < ReceiptWorkOrder.total_count  # 미완료 주문만
                ).order_by(
                    ReceiptWorkOrder.created_at.asc()  # 오래된 것부터 처리
                ).first()

                if order:
                    # CSV 업로드 시에는 URL 중복을 허용 (내용이 다를 수 있으므로)
                    # 추출 시 실제 내용 기반으로 중복 체크
                    existing_review = db.query(Review).join(ReceiptWorkOrder).filter(
                        ReceiptWorkOrder.business_name == business_name,
                        Review.review_url == row['리뷰URL'],
                        Review.content != "내용 추출 대기중"  # 이미 추출된 것만 체크
                    ).count()

                    # 같은 URL이라도 여러 리뷰가 있을 수 있음 (내용이 다른 경우)
                    if existing_review >= 2:  # 이미 2개 이상 추출된 경우만 스킵
                        log.append(f"[SKIP] {business_name} - 동일 URL 리뷰 이미 충분: {row['리뷰URL'][:50]}...")
                        continue

                    # 리뷰를 "내용 추출 대기중" 상태로 생성
                    review_content = "내용 추출 대기중"  # 추후 전체 추출 기능으로 실제 내용 추출

                    review = Review(
                        order_id=order.id,
                        content=review_content,
                        rating=5,
                        review_date=date.today(),
                        review_url=row['리뷰URL'],
                        is_submitted=True,
                        submitted_at=datetime.now()
                    )
                    db.add(review)

                    # CSV/대량 업로드 시에는 카운팅하지 않음
                    # 전체 추출 버튼을 통해 실제 리뷰 내용이 추출되었을 때만 카운팅
                    # order.completed_count += 1 <- 제거

                    count += 1
                    log.append(f"[OK] {business_name} 리뷰 등록 완료 (추출 대기중 - 카운팅 보류)")
                else:
                    log.append(f"[ERROR] 업체명 '{business_name}'에 대한 승인된 미완료 주문을 찾을 수 없음")

            except Exception as e:
                log.append(f"[ERROR] 리뷰 등록 실패: {str(e)}")

        db.commit()
        return {"success": True, "count": count, "log": log, "message": f"{count}개의 리뷰가 등록되었습니다"}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"업로드 실패: {str(e)}", "error": str(e)}

# 템플릿 다운로드
@app.get("/templates/clients_template.xlsx")
async def download_clients_template():
    """고객사 업로드 템플릿 다운로드"""
    data = {
        '회사명': ['예시회사'],
        '담당자': ['홍길동'],
        '단가': [3000],
        '아이디': ['example_id'],
        '비밀번호': ['example_pw']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='고객사등록')

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment; filename=clients_template.xlsx"
        }
    )

@app.get("/templates/reviews_template.xlsx")
async def download_reviews_template():
    """리뷰 업로드 템플릿 다운로드"""
    data = {
        '업체명': ['맛있는집', '카페행복', '스시명가'],
        '리뷰URL': [
            'https://m.place.naver.com/restaurant/1234567/review/123456789',
            'https://m.place.naver.com/restaurant/2345678/review/234567890',
            'https://m.place.naver.com/restaurant/3456789/review/345678901'
        ]
    }

    # 설명 시트용 데이터
    instructions = {
        '항목': ['업체명', '리뷰URL'],
        '설명': [
            '등록된 업체명과 정확히 일치해야 합니다',
            '네이버 플레이스 리뷰 URL을 입력하세요'
        ],
        '예시': [
            '맛있는집',
            'https://m.place.naver.com/restaurant/1234567/review/123456789'
        ]
    }

    df = pd.DataFrame(data)
    df_inst = pd.DataFrame(instructions)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='리뷰등록')
        df_inst.to_excel(writer, index=False, sheet_name='작성방법')

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment; filename=reviews_template.xlsx"
        }
    )

@app.get("/api/admin/reviews-template")
async def download_reviews_template_api():
    """리뷰 업로드 템플릿 다운로드 (API 엔드포인트)"""
    data = {
        '업체명': ['맛있는집', '카페행복', '스시명가'],
        '리뷰URL': [
            'https://m.place.naver.com/restaurant/1234567/review/123456789',
            'https://m.place.naver.com/restaurant/2345678/review/234567890',
            'https://m.place.naver.com/restaurant/3456789/review/345678901'
        ]
    }

    # 설명 시트용 데이터
    instructions = {
        '항목': ['업체명', '리뷰URL'],
        '설명': [
            '등록된 업체명과 정확히 일치해야 합니다',
            '네이버 플레이스 리뷰 URL을 입력하세요'
        ],
        '예시': [
            '맛있는집',
            'https://m.place.naver.com/restaurant/1234567/review/123456789'
        ]
    }

    df = pd.DataFrame(data)
    df_inst = pd.DataFrame(instructions)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='리뷰등록')
        df_inst.to_excel(writer, index=False, sheet_name='작성방법')

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment; filename=reviews_template.xlsx"
        }
    )

@app.get("/templates/bulk_register_template.xlsx")
async def download_bulk_register_template():
    """고객사 대량 업체 등록 템플릿"""
    data = {
        '업체명': ['스타벅스 강남점'],
        '대표자명': ['홍길동'],
        '사업자번호': ['123-45-67890'],
        '플레이스번호': ['12345678'],
        '주소': ['서울시 강남구 테헤란로 123'],
        '작업일수': [5],
        '일일작업량': [10],
        '가이드라인': ['친절하게 작성해주세요']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='업체등록')

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment; filename=bulk_register_template.xlsx"
        }
    )

# 리뷰 관련 API
@app.post("/api/reviews/register")
async def register_review_url(
    request: Request,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    data = await request.json()
    order_id = data.get('order_id')
    review_url = data.get('review_url')

    review_data = extract_naver_review(review_url)

    if review_data:
        new_review = Review(
            order_id=order_id,
            reviewer_id=user.id,
            review_url=review_url,
            content=review_data['content'],
            rating=review_data['rating'],
            review_date=date.today(),
            author_name=review_data['author_name'],
            visit_date=review_data['visit_date'],
            menu_items=json.dumps(review_data['menu_items']),
            images=json.dumps(review_data['images']),
            extracted_at=datetime.now()
        )

        db.add(new_review)

        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()
        if order:
            order.completed_count += 1
            if order.completed_count >= order.total_count:
                order.status = 'completed'
                order.completed_at = datetime.now()

        db.commit()

        return {"success": True, "message": "리뷰가 등록되었습니다", "data": review_data}
    else:
        return {"success": False, "message": "리뷰 추출에 실패했습니다"}

@app.get("/api/reviews/{order_id}")
async def get_order_reviews(
    order_id: int,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    reviews = db.query(Review).filter(Review.order_id == order_id).all()

    result = []
    for review in reviews:
        result.append({
            "id": review.id,
            "content": review.content,
            "rating": review.rating,
            "review_date": review.review_date.isoformat(),
            "author_name": review.author_name,
            "visit_date": review.visit_date,
            "menu_items": json.loads(review.menu_items) if review.menu_items else [],
            "images": json.loads(review.images) if review.images else [],
            "review_url": review.review_url,
            "created_at": review.created_at.isoformat()
        })

    return {"success": True, "data": result}

# 리뷰 삭제 API
@app.delete("/api/reviews/{review_id}")
async def delete_review(
    review_id: int,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    # 관리자 권한 체크
    if user.role not in ["super_admin", "company_admin"]:
        raise HTTPException(status_code=403, detail="권한이 없습니다")

    # 리뷰 찾기
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다")

    # 주문의 완료 카운트 감소
    if review.order_id:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == review.order_id).first()
        if order and order.completed_count > 0:
            order.completed_count -= 1
            # 상태가 완료였다면 진행중으로 변경
            if order.status == 'completed':
                order.status = 'in_progress'
                order.completed_at = None

    # 리뷰 삭제
    db.delete(review)
    db.commit()

    return {"success": True, "message": "리뷰가 삭제되었습니다"}

# 데이터베이스 초기화
def init_db():
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()

    # 기본 회사 생성
    if not db.query(Company).filter(Company.name == "system").first():
        system_company = Company(
            name="system",
            display_name="시스템",
            unit_price=3000
        )
        db.add(system_company)
        db.commit()

    if not db.query(Company).filter(Company.name == "test_company").first():
        test_company = Company(
            name="test_company",
            display_name="테스트 회사",
            unit_price=3500
        )
        db.add(test_company)
        db.commit()

    # 테스트 사용자 생성
    if not db.query(User).filter(User.username == "admin").first():
        admin_user = User(
            username="admin",
            password_hash=hash_password("admin123"),
            full_name="시스템 관리자",
            email="admin@example.com",
            role="super_admin",
            company_id=1,
            commission_rate=0.15
        )
        db.add(admin_user)

    if not db.query(User).filter(User.username == "client").first():
        client_user = User(
            username="client",
            password_hash=hash_password("client123"),
            full_name="고객사 관리자",
            email="client@example.com",
            role="company_admin",
            company_id=2,
            commission_rate=0.1
        )
        db.add(client_user)

    if not db.query(User).filter(User.username == "reviewer").first():
        reviewer_user = User(
            username="reviewer",
            password_hash=hash_password("reviewer123"),
            full_name="리뷰 작성자",
            email="reviewer@example.com",
            role="reviewer",
            company_id=2
        )
        db.add(reviewer_user)

    db.commit()
    db.close()

# 어드민 주문 상세보기 페이지
@app.get("/admin/orders/{order_id}")
async def view_admin_order_details(
    order_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """어드민 주문 상세보기 페이지"""
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다.")

    # 추출 실패한 리뷰를 제외하고 리뷰 목록 조회
    order.reviews = db.query(Review).filter(
        Review.order_id == order_id,
        ~Review.content.like("추출 실패%")
    ).all()

    return templates.TemplateResponse("admin_order_details.html", {
        "request": request,
        "order": order,
        "current_user": user
    })

# 어드민 주문 리포트 다운로드 API
@app.get("/api/admin/orders/{order_id}/download")
async def download_admin_order_report(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """어드민 주문 리포트 다운로드"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

        if not order:
            raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다.")

        # 추출 실패한 리뷰를 제외하고 리뷰 목록 조회
        order.reviews = db.query(Review).filter(
            Review.order_id == order_id,
            ~Review.content.like("추출 실패%")
        ).all()

        # 엑셀 데이터 준비 - 한글 컬럼명 사용, 필요한 정보만 포함
        data = []

        # 주문 기본 정보를 첫 번째 행에 추가
        summary_data = [{
            '업체명': order.business_name,
            '시작일': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
            '종료일': order.end_date.strftime('%Y-%m-%d') if order.end_date else '',
            '일일 개수': order.daily_count,
            '총 일수': order.working_days,
            '총 개수': order.total_count,
            '영수증 날짜': '',
            '리뷰 URL': '',
            '리뷰 내용': ''
        }]

        # 리뷰 데이터 추가 (리뷰 날짜로 정렬)
        if order.reviews:
            # 리뷰를 날짜로 정렬
            sorted_reviews = sorted(order.reviews, key=lambda x: x.review_date if x.review_date else x.created_at)

            for review in sorted_reviews:
                # 모든 리뷰를 data 리스트에 추가
                data.append({
                    '영수증 날짜': review.receipt_date_str or '',
                    '리뷰 URL': review.review_url,
                    '리뷰 내용': review.content or ''
                })

        # 리뷰 데이터만 DataFrame으로 생성 (요약 정보는 별도 처리)
        df = pd.DataFrame(data)
        output = BytesIO()

        # Excel 파일 생성 - xlsxwriter 엔진 사용
        writer = pd.ExcelWriter(output, engine='xlsxwriter')

        # 워크북과 워크시트 가져오기
        workbook = writer.book
        worksheet = workbook.add_worksheet('Report')

        # 셀 포맷 정의
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F0F0F0', 'border': 1})
        info_format = workbook.add_format({'border': 1})

        # A1-A3: 업체 기본 정보를 상단에 배치
        worksheet.write('A1', '업체명:', header_format)
        worksheet.write('B1', summary_data[0]['업체명'], info_format)
        worksheet.write('A2', '작업 기간:', header_format)
        worksheet.write('B2', f"{summary_data[0]['시작일']} ~ {summary_data[0]['종료일']}", info_format)
        worksheet.write('A3', '작업 설정:', header_format)
        worksheet.write('B3', f"일일 {summary_data[0]['일일 개수']}개 × {summary_data[0]['총 일수']}일 = 총 {summary_data[0]['총 개수']}개", info_format)

        # A5부터 리뷰 데이터 테이블 시작
        # 헤더 작성
        headers = ['영수증 날짜', '리뷰 URL', '리뷰 내용']
        for col, header in enumerate(headers):
            worksheet.write(4, col, header, header_format)

        # 리뷰 데이터 작성
        for row_idx, row_data in enumerate(data, start=5):
            worksheet.write(row_idx, 0, row_data.get('영수증 날짜', ''), info_format)
            worksheet.write(row_idx, 1, row_data.get('리뷰 URL', ''), info_format)
            worksheet.write(row_idx, 2, row_data.get('리뷰 내용', ''), info_format)

        # 컬럼 너비 조정
        worksheet.set_column('A:A', 15)  # 영수증 날짜
        worksheet.set_column('B:B', 50)  # 리뷰 URL
        worksheet.set_column('C:C', 80)  # 리뷰 내용

        writer.close()

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 파일명을 안전하게 처리 - 영문만 사용
        filename = f"admin_report_{order.order_no}_{current_time}.xlsx"

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Report generation error: {error_msg}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

# 어드민 전체 리포트 다운로드 API
@app.get("/api/admin/reports/download-all")
async def download_all_admin_reports(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """어드민 전체 리포트 다운로드"""
    try:
        orders = db.query(ReceiptWorkOrder).options(
            joinedload(ReceiptWorkOrder.reviews),
            joinedload(ReceiptWorkOrder.client)
        ).all()

        # 엑셀 데이터 준비 - 모든 필드 포함
        data = []
        for order in orders:
            status_korean = {
                'pending': '미승인',
                'approved': '진행중',
                'completed': '완료',
                'rejected': '거부'
            }.get(order.status, order.status)

            base_info = {
                '주문번호': order.order_no,
                '상호명': order.business_name,
                '대표자명': order.representative_name,
                '사업자번호': order.business_number,
                '플레이스번호': order.place_number if hasattr(order, 'place_number') else '',
                '플레이스링크': order.place_link if hasattr(order, 'place_link') else '',
                '주소': order.business_address if hasattr(order, 'business_address') else '',
                '전화번호': '',  # phone 필드가 User 모델에 있음
                '업종': order.business_type if hasattr(order, 'business_type') else '일반',
                '고객사': order.client.username if order.client else '',
                '접수일': order.created_at.strftime('%Y-%m-%d %H:%M'),
                '시작일': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
                '종료일': order.end_date.strftime('%Y-%m-%d') if order.end_date else '',
                '일일 개수': order.daily_count,
                '총 일수': order.working_days,
                '총 개수': order.total_count,
                '완료 개수': order.completed_count,
                '진행률': f"{(order.completed_count / order.total_count * 100) if order.total_count > 0 else 0:.1f}%",
                '가이드라인': order.guidelines if order.guidelines else '',
                '상태': status_korean
            }
            data.append(base_info)

        df = pd.DataFrame(data)
        output = BytesIO()

        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='All_Orders')
        writer.close()

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"admin_all_reports_{current_time}.xlsx"

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

# 어드민 미승인 리포트 다운로드 API
@app.get("/api/admin/reports/download-pending")
async def download_pending_admin_reports(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """어드민 미승인 리포트 다운로드"""
    try:
        # pending 상태의 주문만 가져오기
        orders = db.query(ReceiptWorkOrder).options(
            joinedload(ReceiptWorkOrder.reviews),
            joinedload(ReceiptWorkOrder.client)
        ).filter(ReceiptWorkOrder.status == 'pending').all()

        # 엑셀 데이터 준비 - 모든 필드 포함
        data = []
        for order in orders:
            data.append({
                '주문번호': order.order_no,
                '상호명': order.business_name,
                '대표자명': order.representative_name,
                '사업자번호': order.business_number,
                '플레이스번호': order.place_number if hasattr(order, 'place_number') else '',
                '플레이스링크': order.place_link if hasattr(order, 'place_link') else '',
                '주소': order.business_address if hasattr(order, 'business_address') else '',
                '전화번호': '',  # phone 필드가 User 모델에 있음
                '업종': order.business_type if hasattr(order, 'business_type') else '일반',
                '고객사': order.client.username if order.client else '',
                '접수일': order.created_at.strftime('%Y-%m-%d %H:%M'),
                '시작일': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
                '종료일': order.end_date.strftime('%Y-%m-%d') if order.end_date else '',
                '일일 개수': order.daily_count,
                '총 일수': order.working_days,
                '총 개수': order.total_count,
                '가이드라인': order.guidelines if order.guidelines else '',
                '상태': '미승인'
            })

        if not data:
            # 미승인 주문이 없을 경우 빈 데이터프레임 생성
            data = [{'알림': '미승인 주문이 없습니다.'}]

        df = pd.DataFrame(data)
        output = BytesIO()

        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Pending_Orders')

        # 워크북과 워크시트 가져오기
        workbook = writer.book
        worksheet = writer.sheets['Pending_Orders']

        # 헤더 포맷 정의
        header_format = workbook.add_format({'bold': True, 'bg_color': '#FFF2CC', 'border': 1})

        # 헤더 스타일 적용
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # 컬럼 너비 자동 조정
        for i, col in enumerate(df.columns):
            column_width = max(df[col].astype(str).str.len().max(), len(str(col))) + 2
            worksheet.set_column(i, i, min(column_width, 30))

        writer.close()

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"admin_pending_reports_{current_time}.xlsx"

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

# 개별 주문 리포트 다운로드 API
@app.get("/api/client/report/download/{order_id}")
async def download_order_report(
    order_id: int,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """개별 주문 리포트 다운로드"""
    try:
        order = db.query(ReceiptWorkOrder).options(
            joinedload(ReceiptWorkOrder.reviews)
        ).filter(
            ReceiptWorkOrder.id == order_id,
            ReceiptWorkOrder.client_id == user.id
        ).first()

        if not order:
            raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다.")

        # 엑셀 데이터 준비 - 한글 컬럼명 사용, 필요한 정보만 포함
        data = []

        # 주문 기본 정보를 첫 번째 행에 추가
        summary_data = [{
            '업체명': order.business_name,
            '시작일': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
            '종료일': order.end_date.strftime('%Y-%m-%d') if order.end_date else '',
            '일일 개수': order.daily_count,
            '총 일수': order.working_days,
            '총 개수': order.total_count,
            '영수증 날짜': '',
            '리뷰 URL': '',
            '리뷰 내용': ''
        }]

        # 리뷰 데이터 추가 (리뷰 날짜로 정렬)
        if order.reviews:
            # 리뷰를 날짜로 정렬
            sorted_reviews = sorted(order.reviews, key=lambda x: x.review_date if x.review_date else x.created_at)

            for review in sorted_reviews:
                # 모든 리뷰를 data 리스트에 추가
                data.append({
                    '영수증 날짜': review.receipt_date_str or '',
                    '리뷰 URL': review.review_url,
                    '리뷰 내용': review.content or ''
                })

        # 리뷰 데이터만 DataFrame으로 생성 (요약 정보는 별도 처리)
        df = pd.DataFrame(data)
        output = BytesIO()

        # Excel 파일 생성 - xlsxwriter 엔진 사용
        writer = pd.ExcelWriter(output, engine='xlsxwriter')

        # 워크북과 워크시트 가져오기
        workbook = writer.book
        worksheet = workbook.add_worksheet('Report')

        # 셀 포맷 정의
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F0F0F0', 'border': 1})
        info_format = workbook.add_format({'border': 1})

        # A1-A3: 업체 기본 정보를 상단에 배치
        worksheet.write('A1', '업체명:', header_format)
        worksheet.write('B1', summary_data[0]['업체명'], info_format)
        worksheet.write('A2', '작업 기간:', header_format)
        worksheet.write('B2', f"{summary_data[0]['시작일']} ~ {summary_data[0]['종료일']}", info_format)
        worksheet.write('A3', '작업 설정:', header_format)
        worksheet.write('B3', f"일일 {summary_data[0]['일일 개수']}개 × {summary_data[0]['총 일수']}일 = 총 {summary_data[0]['총 개수']}개", info_format)

        # A5부터 리뷰 데이터 테이블 시작
        # 헤더 작성
        headers = ['영수증 날짜', '리뷰 URL', '리뷰 내용']
        for col, header in enumerate(headers):
            worksheet.write(4, col, header, header_format)

        # 리뷰 데이터 작성
        for row_idx, row_data in enumerate(data, start=5):
            worksheet.write(row_idx, 0, row_data.get('영수증 날짜', ''), info_format)
            worksheet.write(row_idx, 1, row_data.get('리뷰 URL', ''), info_format)
            worksheet.write(row_idx, 2, row_data.get('리뷰 내용', ''), info_format)

        # 컬럼 너비 조정
        worksheet.set_column('A:A', 15)  # 영수증 날짜
        worksheet.set_column('B:B', 50)  # 리뷰 URL
        worksheet.set_column('C:C', 80)  # 리뷰 내용

        writer.close()

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 파일명을 안전하게 처리 - 영문만 사용
        filename = f"report_{order.order_no}_{current_time}.xlsx"

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Report generation error: {error_msg}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

# 전체 리포트 다운로드 API
@app.get("/api/client/report/download")
async def download_client_report(
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """고객사 리포트 다운로드"""
    try:
        orders = db.query(ReceiptWorkOrder).options(
            joinedload(ReceiptWorkOrder.reviews)
        ).filter(ReceiptWorkOrder.client_id == user.id).all()

        # 엑셀 데이터 준비 - 영문 컬럼명 사용
        data = []
        for order in orders:
            base_info = {
                'Order_No': order.order_no,
                'Business_Name': order.business_name,
                'Request_Date': order.created_at.strftime('%Y-%m-%d'),
                'Start_Date': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
                'Status': order.status,
                'Total_Count': order.total_count,
                'Completed_Count': order.completed_count,
                'Progress_Rate': f"{(order.completed_count / order.total_count * 100) if order.total_count > 0 else 0:.1f}%"
            }

            if order.reviews:
                for review in order.reviews:
                    data.append({
                        **base_info,
                        'Review_Date': review.review_date.strftime('%Y-%m-%d'),
                        'Review_URL': review.review_url,
                        'Review_Content': review.content or '',
                        'Rating': review.rating
                    })
            else:
                data.append({
                    **base_info,
                    'Review_Date': '',
                    'Review_URL': '',
                    'Review_Content': '',
                    'Rating': ''
                })

        df = pd.DataFrame(data)
        output = BytesIO()

        # Excel 파일 생성 - xlsxwriter 엔진 사용
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Report')
        writer.close()

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 파일명을 안전하게 처리 - 영문만 사용
        filename = f"report_{user.username}_{current_time}.xlsx"

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Report generation error: {error_msg}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

# 연장 신청 API
@app.post("/api/client/extension/request")
async def request_extension(
    request: Request,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """연장 신청"""
    try:
        form = await request.json()
        daily_count = form.get('daily_count')
        total_days = form.get('total_days')

        if not daily_count or not total_days:
            return {"success": False, "message": "필수 정보가 누락되었습니다."}

        extension = ExtensionRequest(
            client_id=user.id,
            daily_count=daily_count,
            total_days=total_days,
            status='pending'
        )

        db.add(extension)
        db.commit()

        return {"success": True, "message": "연장 신청이 완료되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"연장 신청 실패: {str(e)}"}

# 리뷰 내용 추출 API (복구)
@app.post("/api/admin/extract-review-content")
async def extract_review_content(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """리뷰 내용 추출"""
    try:
        form = await request.json()
        review_url = form.get('review_url', '').strip()

        if not review_url:
            return {"success": False, "message": "리뷰 URL을 입력해주세요."}

        # 간단한 리뷰 내용 추출 (실제로는 웹 스크래핑 등을 사용)
        extracted_content = {
            "content": "리뷰 내용 추출 완료",
            "rating": 5,
            "date": datetime.now().strftime('%Y-%m-%d'),
            "url": review_url
        }

        if 'naver.com' in review_url:
            extracted_content["content"] = "네이버 리뷰 - 맛있게 잘 먹었습니다!"
        elif 'kakao.com' in review_url:
            extracted_content["content"] = "카카오맵 리뷰 - 좋은 서비스였습니다!"

        return {"success": True, "data": extracted_content}

    except Exception as e:
        return {"success": False, "message": f"리뷰 내용 추출 실패: {str(e)}"}

# 인센티브 캘린더 API
@app.get("/api/admin/incentive/calendar")
async def get_incentive_calendar(
    user_id: int,
    start_date: str,
    end_date: str,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """인센티브 캘린더 데이터 조회"""
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        manager = db.query(Manager).filter(Manager.id == user_id).first()
        if not manager:
            return {"success": False, "message": "담당자를 찾을 수 없습니다."}

        # 해당 기간 완료된 작업 조회 (모든 완료된 작업에서 담당자별로 인센티브 계산)
        completed_orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.status == 'completed',
            ReceiptWorkOrder.completed_at.between(start, end)
        ).all()

        total_revenue = 0
        daily_stats = {}

        for order in completed_orders:
            date_key = order.completed_at.strftime('%Y-%m-%d')
            if date_key not in daily_stats:
                daily_stats[date_key] = {
                    'count': 0,
                    'revenue': 0
                }

            revenue = order.total_count * (order.company.unit_price if order.company else 3000)
            daily_stats[date_key]['count'] += order.total_count
            daily_stats[date_key]['revenue'] += revenue
            total_revenue += revenue

        incentive = total_revenue * manager.commission_rate

        return {
            "success": True,
            "data": {
                "user": {
                    "id": manager.id,
                    "name": manager.name,
                    "commission_rate": manager.commission_rate
                },
                "period": {
                    "start": start_date,
                    "end": end_date
                },
                "summary": {
                    "total_revenue": total_revenue,
                    "total_incentive": incentive,
                    "total_orders": len(completed_orders)
                },
                "daily_stats": daily_stats
            }
        }

    except Exception as e:
        return {"success": False, "message": f"데이터 조회 실패: {str(e)}"}

# 연장 신청 관련 API들
@app.get("/api/admin/extensions")
async def get_extensions(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """연장 신청 목록 조회"""
    try:
        extensions = db.query(ExtensionRequest).options(
            joinedload(ExtensionRequest.client)
        ).order_by(ExtensionRequest.created_at.desc()).all()

        result = []
        for ext in extensions:
            result.append({
                "id": ext.id,
                "order_id": ext.order_id,
                "client_id": ext.client_id,
                "client_username": ext.client.username if ext.client else "알 수 없음",
                "daily_count": ext.daily_count,
                "total_days": ext.total_days,
                "status": ext.status,
                "created_at": ext.created_at.strftime('%Y-%m-%d %H:%M'),
                "processed_at": ext.processed_at.strftime('%Y-%m-%d %H:%M') if ext.processed_at else ""
            })

        return result

    except Exception as e:
        return {"success": False, "message": str(e)}

@app.put("/api/admin/receipt/order/{order_id}/memo")
async def update_order_memo(
    order_id: int,
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """주문 메모 업데이트"""
    try:
        data = await request.json()
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

        if not order:
            return {"success": False, "message": "주문을 찾을 수 없습니다."}

        order.admin_memo = data.get('memo', '')
        db.commit()

        return {"success": True, "message": "메모가 저장되었습니다."}

    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/admin/extensions/{ext_id}/approve")
async def approve_extension(
    ext_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """연장 신청 승인"""
    try:
        extension = db.query(ExtensionRequest).filter(ExtensionRequest.id == ext_id).first()
        if not extension:
            return {"success": False, "message": "연장 신청을 찾을 수 없습니다."}

        extension.status = 'approved'
        extension.processed_at = datetime.now()
        extension.processed_by = user.id

        # 실제 주문 연장 처리
        order = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.id == extension.order_id
        ).first()

        if order:
            # 기존 종료일 기준으로 연장
            new_end_date = order.end_date + timedelta(days=extension.total_days)
            order.end_date = new_end_date
            order.working_days += extension.total_days
            order.total_count += (extension.daily_count * extension.total_days)

            # 일일 작업량도 업데이트 (평균값으로)
            if extension.daily_count != order.daily_count:
                order.daily_count = extension.daily_count

        db.commit()
        return {"success": True, "message": "연장 신청이 승인되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

@app.post("/api/admin/extensions/{ext_id}/reject")
async def reject_extension(
    ext_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """연장 신청 거부"""
    try:
        extension = db.query(ExtensionRequest).filter(ExtensionRequest.id == ext_id).first()
        if not extension:
            return {"success": False, "message": "연장 신청을 찾을 수 없습니다."}

        extension.status = 'rejected'
        extension.processed_at = datetime.now()
        extension.processed_by = user.id

        db.commit()
        return {"success": True, "message": "연장 신청이 거부되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

# 미승인 작업 엑셀 다운로드 API
@app.get("/api/admin/export/pending-orders")
async def export_pending_orders(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """미승인 작업 엑셀 다운로드"""
    try:
        orders = db.query(ReceiptWorkOrder).options(
            joinedload(ReceiptWorkOrder.company),
            joinedload(ReceiptWorkOrder.client)
        ).filter(ReceiptWorkOrder.status == 'pending').all()

        # 엑셀 데이터 준비
        data = []
        for order in orders:
            # 고객사 정보
            company_name = order.company.display_name if order.company else "알 수 없음"

            data.append({
                '고객사': company_name,
                '상호명': order.business_name,
                '대표자명': order.representative_name,
                '사업자번호': order.business_number,
                '업종': order.business_type,
                '플레이스 번호': order.place_number if order.place_number else '',
                '플레이스 링크': order.place_link if order.place_link else '',
                '주소': order.business_address,
                '접수일': order.created_at.strftime('%Y-%m-%d'),
                '시작일': order.start_date.strftime('%Y-%m-%d') if order.start_date else '',
                '하루 발행수': order.daily_count,
                '작업일수': order.working_days,
                '종료일': order.end_date.strftime('%Y-%m-%d') if order.end_date else '',
                '총 갯수': order.total_count,
                '건당 금액': order.unit_price,
                '주문 금액': order.total_price
            })

        # 총계 계산
        total_count = sum(order.total_count for order in orders) if orders else 0
        total_amount = sum(order.total_price for order in orders) if orders else 0

        # 데이터가 없어도 빈 엑셀 파일 생성
        if not data:
            data = [{
                '고객사': '',
                '상호명': '',
                '대표자명': '',
                '사업자번호': '',
                '업종': '',
                '플레이스 번호': '',
                '플레이스 링크': '',
                '주소': '',
                '접수일': '',
                '시작일': '',
                '하루 발행수': '',
                '작업일수': '',
                '종료일': '',
                '총 갯수': '',
                '건당 금액': '',
                '주문 금액': ''
            }]
        else:
            # 총계 행 추가
            data.append({
                '고객사': '총계',
                '상호명': '',
                '대표자명': '',
                '사업자번호': '',
                '업종': '',
                '플레이스 번호': '',
                '플레이스 링크': '',
                '주소': '',
                '접수일': '',
                '시작일': '',
                '하루 발행수': '',
                '작업일수': '',
                '종료일': '',
                '총 갯수': total_count,
                '건당 금액': '',
                '주문 금액': total_amount
            })

        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='미승인작업')

            # 워크시트와 워크북 가져오기
            workbook = writer.book
            worksheet = writer.sheets['미승인작업']

            # 총계 행 스타일링 (마지막 행이 총계)
            if len(data) > 1:  # 데이터가 있을 때만
                total_row = len(data) + 1  # 헤더 포함해서 +1

                # 총계 행 배경색과 굵은 글씨 적용
                from openpyxl.styles import Font, PatternFill
                bold_font = Font(bold=True)
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=total_row, column=col)
                    cell.font = bold_font
                    cell.fill = yellow_fill

        output.seek(0)
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename=pending_orders_{current_time}.xlsx"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 생성 실패: {str(e)}")

# 고객사 추가 API
@app.post("/api/admin/clients")
async def add_client(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """고객사 추가"""
    try:
        form = await request.json()

        # 중복 체크
        existing = db.query(Company).filter(Company.name == form['name']).first()
        if existing:
            return {"success": False, "message": "이미 존재하는 회사명입니다."}

        company = Company(
            name=form['name'],
            display_name=form['display_name'],
            unit_price=float(form.get('unit_price', 3000))
        )

        db.add(company)
        db.commit()

        return {"success": True, "message": "고객사가 추가되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"추가 실패: {str(e)}"}

# 리뷰어 추가 API
@app.post("/api/admin/reviewers")
async def add_reviewer(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """리뷰어 추가"""
    try:
        form = await request.json()

        # 중복 체크
        existing = db.query(User).filter(User.username == form['username']).first()
        if existing:
            return {"success": False, "message": "이미 존재하는 사용자명입니다."}

        # 패스워드 해시화
        password_hash = hashlib.sha256(form['password'].encode()).hexdigest()

        reviewer = User(
            username=form['username'],
            password_hash=password_hash,
            full_name=form.get('full_name', ''),
            role='reviewer',
            commission_rate=float(form.get('commission_rate', 0.1)),  # 이미 소수점으로 받음
            is_active=True
        )

        db.add(reviewer)
        db.commit()

        return {"success": True, "message": "리뷰어가 추가되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"추가 실패: {str(e)}"}

# 리뷰어 추가 API (별도 경로)
@app.post("/api/admin/reviewers/add")
async def add_reviewer_alt(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """리뷰어 추가 (대체 경로)"""
    return await add_reviewer(request, user, db)

# 전체 리뷰 목록 조회 API
@app.get("/api/admin/reviews")
async def get_all_reviews(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """전체 리뷰 목록 조회"""
    try:
        reviews = db.query(Review).options(
            joinedload(Review.order).joinedload(ReceiptWorkOrder.company)
        ).order_by(Review.created_at.desc()).all()

        result = []
        for review in reviews:
            result.append({
                "id": review.id,
                "order_no": review.order.order_no if review.order else "알 수 없음",
                "company_name": review.order.company.display_name if review.order and review.order.company else "알 수 없음",
                "business_name": review.order.business_name if review.order else "알 수 없음",
                "review_url": review.review_url,
                "content": review.content or "",
                "rating": review.rating,
                "review_date": review.review_date.strftime('%Y-%m-%d'),
                "receipt_date": review.receipt_date_str or "",  # 영수증 날짜 추가
                "is_submitted": review.is_submitted,
                "created_at": review.created_at.strftime('%Y-%m-%d %H:%M')
            })

        return result

    except Exception as e:
        return {"success": False, "message": str(e)}

# 대량 리뷰 업로드 API
@app.post("/api/admin/bulk-upload-reviews")
async def bulk_upload_reviews(
    file: UploadFile,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """대량 리뷰 업로드"""
    try:
        # 파일 읽기
        content = await file.read()
        df = read_file_to_dataframe(content, file.filename)

        # 디버깅: 실제 컬럼 확인
        actual_columns = list(df.columns)

        # 필수 컬럼 확인
        required_columns = ['업체명', '리뷰URL']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return {
                "success": False,
                "message": f"필수 컬럼이 누락되었습니다. 필요: {', '.join(required_columns)}, 실제: {', '.join(actual_columns)}"
            }

        # 데이터가 있는지 확인
        if df.empty:
            return {"success": False, "message": "업로드한 파일에 데이터가 없습니다."}

        # 리뷰 추출기 초기화 시도
        try:
            from real_review_extractor import get_extractor
            extractor = get_extractor()
            use_extractor = True
            print("[OK] 리뷰 추출기 초기화 성공")
        except Exception as e:
            print(f"[WARNING] 리뷰 추출기 초기화 실패: {e}")
            extractor = None
            use_extractor = False

        log = []
        success_count = 0
        total_rows = len(df)

        for idx, row in df.iterrows():
            try:
                # 데이터 정리 - nan 처리 및 공백 제거
                business_name = str(row['업체명']).strip() if pd.notna(row['업체명']) else ''
                review_url = str(row['리뷰URL']).strip() if pd.notna(row['리뷰URL']) else ''

                # URL 정리 - 백슬래시, 줄바꿈 등 제거
                review_url = review_url.replace('\\', '').replace('\n', '').replace('\r', '').strip()

                # 디버깅 로그
                print(f"행 {idx+1} - 업체명: '{business_name}', URL: '{review_url}'")

                if not business_name or not review_url:
                    log.append(f"행 {idx+1}: 필수 정보 누락 (업체명: '{business_name}', URL: '{review_url}')")
                    continue

                # CSV 업로드 시에는 URL 중복을 허용 (내용이 다를 수 있으므로)
                # 추출 시 실제 내용 기반으로 중복 체크
                existing_count = db.query(Review).join(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.business_name == business_name,
                    Review.review_url == review_url,
                    Review.content != "내용 추출 대기중"  # 이미 추출된 것만 체크
                ).count()

                # 같은 URL이라도 여러 리뷰가 있을 수 있음 (내용이 다른 경우)
                if existing_count >= 2:  # 이미 2개 이상 추출된 경우만 스킵
                    log.append(f"행 {idx+1}: [SKIP] 동일 URL 리뷰 이미 충분 - {business_name}")
                    continue

                # 스마트 자동 배정 시도
                from smart_auto_assign import smart_assign_review, extract_shop_name_from_content

                # CSV에 업체명이 제공되지 않았거나 "자동매칭"인 경우
                if not business_name or business_name == "자동매칭":
                    # 리뷰 내용에서 업체명 추출 시도
                    if review_content and review_content != "내용 추출 대기중":
                        detected_shop = extract_shop_name_from_content(review_content)
                        if detected_shop:
                            business_name = detected_shop
                            log.append(f"행 {idx+1}: 리뷰에서 업체명 '{business_name}' 자동 감지")

                # 해당 업체명과 매칭되는 미완료 주문 찾기 (가장 최근 것부터 - 연장 주문 우선)
                order = db.query(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.business_name == business_name,
                    ReceiptWorkOrder.status == 'approved',
                    ReceiptWorkOrder.completed_count < ReceiptWorkOrder.total_count  # 미완료 주문만
                ).order_by(
                    ReceiptWorkOrder.created_at.asc()  # 오래된 것부터 처리 (순차적으로)
                ).first()

                # 정확히 일치하는 주문이 없으면 부분 일치 시도
                if not order and business_name:
                    order = db.query(ReceiptWorkOrder).filter(
                        ReceiptWorkOrder.business_name.contains(business_name),
                        ReceiptWorkOrder.status == 'approved',
                        ReceiptWorkOrder.completed_count < ReceiptWorkOrder.total_count
                    ).order_by(
                        ReceiptWorkOrder.created_at.asc()  # 오래된 것부터 처리 (순차적으로)
                    ).first()

                    if order:
                        log.append(f"행 {idx+1}: 부분 일치로 '{order.business_name}' 주문에 배정")

                if not order:
                    # 승인된 모든 주문의 업체명 확인
                    all_approved = db.query(ReceiptWorkOrder.business_name).filter(
                        ReceiptWorkOrder.status == 'approved'
                    ).all()
                    approved_names = [o[0] for o in all_approved]

                    # 완료된 주문이 있는지도 확인
                    completed_exists = db.query(ReceiptWorkOrder).filter(
                        ReceiptWorkOrder.business_name == business_name,
                        ReceiptWorkOrder.status == 'completed'
                    ).first()

                    if completed_exists:
                        log.append(f"행 {idx+1}: '{business_name}' 모든 주문이 완료됨")
                    else:
                        log.append(f"행 {idx+1}: '{business_name}'과 매칭되는 승인된 주문이 없음. 승인된 업체: {approved_names}")
                    continue

                # 리뷰 내용과 날짜 추출
                review_content = None
                receipt_date_str = None

                # CSV 업로드 시에는 추출하지 않음 (전체 추출 버튼으로만 추출)
                # 항상 추출 대기 상태로 설정
                review_content = "내용 추출 대기중"
                receipt_date_str = None

                # 중복 체크 - 동일 URL과 동일 내용이면 카운팅 안함
                should_count = True
                skip_creation = False

                if existing:
                    # 동일 URL과 동일 내용이면
                    if existing.content == review_content:
                        should_count = False  # 카운팅 안함

                        # 같은 주문이면 완전 스킵
                        if existing.order_id == order.id:
                            log.append(f"행 {idx+1}: 완전 중복 (동일 주문, URL, 내용) - 스킵")
                            skip_creation = True
                        else:
                            # 다른 주문에 동일 URL/내용이면 스킵 (등록하지 않음)
                            log.append(f"행 {idx+1}: 다른 주문에 동일 URL/내용 존재 - 스킵")
                            skip_creation = True
                    else:
                        # URL은 같지만 내용이 다르면 새 리뷰로 처리
                        log.append(f"행 {idx+1}: URL은 같지만 내용이 다름 - 새 리뷰로 등록")
                        should_count = True

                # 스킵해야 하는 경우 다음 행으로
                if skip_creation:
                    continue

                # 리뷰 생성
                review = Review(
                    order_id=order.id,
                    review_url=review_url,
                    review_date=datetime.now().date(),
                    content=review_content,
                    receipt_date_str=receipt_date_str,  # 영수증 날짜 저장
                    rating=int(row.get('평점', 5)) if pd.notna(row.get('평점')) else 5,
                    is_submitted=True,
                    created_at=datetime.now()
                )

                db.add(review)

                # CSV 업로드 시에는 절대 카운팅하지 않음
                # 전체 추출 버튼을 통해서만 카운팅됨
                log.append(f"행 {idx+1}: 리뷰 등록 (추출 대기중 - 카운팅 보류)")

                db.commit()

                log.append(f"[OK] 등록완료: {business_name} (리뷰 {len(review_content)}자)")
                success_count += 1

            except Exception as e:
                log.append(f"행 {idx+1}: 오류 - {str(e)}")
                db.rollback()
                continue

        # 상세한 결과 메시지 생성
        if total_rows > 0:
            message = f"전체 {total_rows}행 중 {success_count}건 등록완료"
            if success_count == 0 and log:
                # 모두 실패한 경우 상세 이유 표시
                message += f"\n\n실패 내역:\n" + "\n".join(log[:5])  # 처음 5개만
                if len(log) > 5:
                    message += f"\n... 외 {len(log) - 5}건"
        else:
            message = "데이터가 없습니다"

        return {
            "success": True,
            "message": message,
            "count": success_count,
            "log": log
        }

    except Exception as e:
        return {"success": False, "message": f"파일 처리 실패: {str(e)}"}


# 담당자 추가 API
@app.post("/api/admin/managers/add")
async def add_manager(
    request: Request,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """담당자 추가"""
    try:
        form = await request.json()

        # 중복 체크
        existing = db.query(Manager).filter(Manager.name == form['name']).first()
        if existing:
            return {"success": False, "message": "이미 존재하는 담당자명입니다."}

        manager = Manager(
            name=form['name'],
            commission_rate=float(form.get('commission_rate', 0.1))
        )

        db.add(manager)
        db.commit()

        return {"success": True, "message": "담당자가 추가되었습니다."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"추가 실패: {str(e)}"}

# 담당자 목록 조회 API
@app.get("/api/admin/managers")
async def get_managers(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """담당자 목록 조회"""
    try:
        managers = db.query(Manager).filter(Manager.is_active == True).all()

        result = []
        for manager in managers:
            result.append({
                "id": manager.id,
                "name": manager.name,
                "commission_rate": manager.commission_rate,
                "created_at": manager.created_at.strftime('%Y-%m-%d') if manager.created_at else ""
            })

        return result

    except Exception as e:
        return {"success": False, "message": str(e)}

if __name__ == "__main__":
    init_db()
    print("=" * 70)
    print("네이버 리뷰 관리 시스템 - 최종 완성판")
    print("=" * 70)
    print("서버 시작: http://localhost:8000")
    print("")
    print("테스트 계정:")
    print("  시스템 관리자: admin / admin123")
    print("  고객사 관리자: client / client123")
    print("  리뷰 작성자: reviewer / reviewer123")
    print("")
    print("새로운 기능:")
    print("  - CSV, Excel, 한컴 파일 지원")
    print("  - 고객사별 단가 설정")
    print("  - 고객사 대량 업체 등록")
    print("  - 인센티브 대시보드")
    print("  - 미승인 작업만 엑셀 다운로드")
    print("  - 템플릿 다운로드")
    print("=" * 70)

# 리뷰 삭제 API
# 주의: delete-all 라우트를 {review_id} 라우트보다 먼저 정의해야 함
@app.delete("/api/admin/reviews/delete-all")
async def delete_all_reviews(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """모든 리뷰 삭제"""
    try:
        count = db.query(Review).count()

        # 모든 주문의 완료 카운트 리셋
        orders = db.query(ReceiptWorkOrder).all()
        for order in orders:
            order.completed_count = 0
            if order.status == 'completed':
                order.status = 'approved'
                order.completed_at = None

        # 모든 리뷰 삭제
        db.query(Review).delete()
        db.commit()

        return {"success": True, "message": f"{count}개의 리뷰가 삭제되었습니다"}
    except Exception as e:
        db.rollback()
        print(f"전체 삭제 오류: {e}")
        return {"success": False, "message": f"삭제 중 오류 발생: {str(e)}"}

@app.delete("/api/admin/reviews/{review_id}")
async def delete_review(
    review_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """개별 리뷰 삭제"""
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다")

    # 주문의 완료 카운트 감소
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == review.order_id).first()
    if order:
        order.completed_count = max(0, order.completed_count - 1)
        if order.status == 'completed':
            order.status = 'approved'

    db.delete(review)
    db.commit()
    return {"success": True, "message": "리뷰가 삭제되었습니다"}

# 리뷰 자동 재배정 함수
def auto_redistribute_reviews(db: Session):
    """완료된 주문의 초과 리뷰를 다음 주문으로 자동 이관"""
    try:
        # 초과 리뷰가 있는 완료된 주문 찾기
        completed_orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.completed_count >= ReceiptWorkOrder.total_count
        ).all()

        for order in completed_orders:
            # 해당 주문의 모든 리뷰 가져오기
            reviews = db.query(Review).filter(
                Review.order_id == order.id,
                Review.content != "내용 추출 대기중"
            ).all()

            if len(reviews) > order.total_count:
                # 초과 리뷰가 있는 경우
                excess_count = len(reviews) - order.total_count
                excess_reviews = reviews[order.total_count:]  # 초과분만 선택

                # 같은 업체명의 다음 주문 찾기
                next_order = db.query(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.business_name == order.business_name,
                    ReceiptWorkOrder.id != order.id,
                    ReceiptWorkOrder.completed_count < ReceiptWorkOrder.total_count,
                    ReceiptWorkOrder.status == 'approved'
                ).order_by(
                    ReceiptWorkOrder.created_at.asc()  # 오래된 것부터
                ).first()

                if next_order:
                    # 초과 리뷰를 다음 주문으로 이관
                    transferred = 0
                    for review in excess_reviews:
                        if next_order.completed_count < next_order.total_count:
                            review.order_id = next_order.id
                            next_order.completed_count += 1
                            # 완료된 주문의 카운트는 변경하지 않음 (total_count 유지)
                            # order.completed_count를 감소시키지 않음
                            transferred += 1
                            print(f"[자동이관] 리뷰 {review.id} -> 주문 {next_order.id} ({next_order.business_name})")
                        else:
                            break

                    if transferred > 0:
                        # 완료된 주문은 total_count를 유지
                        # 이관 후에도 완료 상태 및 카운트 유지
                        print(f"[자동이관] {order.business_name}: {transferred}개 리뷰 이관 완료")
                        db.commit()
    except Exception as e:
        print(f"[자동이관] 오류: {e}")
        db.rollback()

# 백그라운드 추출 작업 함수
def extract_reviews_background(review_ids: List[int]):
    """백그라운드에서 리뷰 내용 추출"""
    print(f"[백그라운드] 추출 시작: {len(review_ids)}개 리뷰")
    db = SessionLocal()
    try:
        # 먼저 기존의 "추출 실패" 리뷰들을 모두 삭제
        failed_reviews = db.query(Review).filter(
            Review.content.like("추출 실패%")
        ).all()
        if failed_reviews:
            print(f"[백그라운드] 기존 추출 실패 리뷰 {len(failed_reviews)}개 삭제 중...")
            for failed_review in failed_reviews:
                db.delete(failed_review)
            db.commit()
            print(f"[백그라운드] 추출 실패 리뷰 삭제 완료")

        from real_review_extractor import get_extractor
        extractor = get_extractor()
        print(f"[백그라운드] extractor 생성 완료")

        for review_id in review_ids:
            print(f"[백그라운드] 리뷰 {review_id} 추출 시작")
            try:
                review = db.query(Review).filter(Review.id == review_id).first()
                if not review:
                    continue

                # 업체명 가져오기
                order = db.query(ReceiptWorkOrder).filter(
                    ReceiptWorkOrder.id == review.order_id
                ).first()

                shop_name = order.business_name if order else None

                # URL에서 리뷰 내용 추출 (재시도 로직 포함)
                max_retries = 2
                review_text = None
                receipt_date = None
                metadata = {}

                for attempt in range(max_retries):
                    try:
                        print(f"[시도 {attempt + 1}/{max_retries}] 리뷰 {review_id} 추출 중...")
                        review_text, receipt_date, metadata = extractor.extract_review(
                            review.review_url,
                            shop_name
                        )

                        # 성공했으면 중단
                        if review_text and "오류" not in review_text and "찾을 수 없습니다" not in review_text:
                            print(f"[성공] 리뷰 {review_id} 추출 성공 (시도 {attempt + 1})")
                            break
                        else:
                            # 실패했고 재시도 남아있으면 잠시 대기
                            if attempt < max_retries - 1:
                                print(f"[재시도] 리뷰 {review_id} 추출 실패, 3초 후 재시도...")
                                import time
                                time.sleep(3)
                    except Exception as retry_error:
                        print(f"[오류] 리뷰 {review_id} 추출 중 오류 (시도 {attempt + 1}): {retry_error}")
                        if attempt < max_retries - 1:
                            import time
                            time.sleep(3)

                if review_text and "오류" not in review_text and "찾을 수 없습니다" not in review_text:
                    # 이전 내용 저장 (카운팅 판단용)
                    old_content = review.content

                    review.content = review_text
                    review.receipt_date_str = receipt_date
                    review.extracted_at = datetime.now()

                    # 추출 성공 시 항상 주문 상태 확인 및 업데이트
                    if order:
                        # 처음 추출 성공한 경우에만 카운팅
                        if old_content == "내용 추출 대기중":
                            # 중복 체크 - 같은 업체명, 같은 URL, 같은 내용이 이미 추출된 경우
                            existing = db.query(Review).join(ReceiptWorkOrder).filter(
                                Review.id != review.id,
                                ReceiptWorkOrder.business_name == order.business_name,
                                Review.review_url == review.review_url,
                                Review.content == review_text,  # 같은 내용
                                Review.content != "내용 추출 대기중",
                                Review.content != None,
                                Review.content != ""
                            ).first()

                            if existing:
                                # 중복이면 현재 리뷰 삭제 (고객에게 보이지 않도록)
                                print(f"[중복 발견] 리뷰 {review_id} - 동일 업체명/URL/내용 (삭제)")
                                db.delete(review)
                                db.commit()
                                continue  # 다음 리뷰로
                            else:
                                # 중복이 아닌 경우만 카운팅
                                if order.completed_count < order.total_count:
                                    order.completed_count += 1
                                    print(f"리뷰 {review_id} 추출 완료 - 카운팅: {order.completed_count}/{order.total_count}")
                                else:
                                    print(f"리뷰 {review_id} - 이미 목표 달성 (스킵): {order.completed_count}/{order.total_count}")

                        # 추출이 완료된 리뷰 수를 다시 확인하고 주문 상태 업데이트
                        extracted_count = db.query(Review).filter(
                            Review.order_id == order.id,
                            Review.content != "내용 추출 대기중",
                            Review.content != None,
                            Review.content != "",
                            ~Review.content.like("추출 실패%")
                        ).count()

                        # 추출된 리뷰 수가 목표에 도달하면 completed로 변경
                        if extracted_count >= order.total_count and order.status != 'completed':
                            order.status = 'completed'
                            order.completed_at = datetime.now()
                            order.completed_count = extracted_count
                            print(f"[주문 완료] 주문 {order.id} - {order.business_name}: {extracted_count}/{order.total_count} 리뷰 추출 완료")
                else:
                    # 추출 실패한 리뷰는 삭제
                    print(f"리뷰 {review_id} 추출 실패: {review_text[:50] if review_text else '알 수 없음'} - 리뷰 삭제")
                    db.delete(review)

                db.commit()

                # 다음 리뷰 추출 전 잠시 대기 (브라우저 안정화)
                import time
                time.sleep(1)

            except Exception as e:
                print(f"리뷰 {review_id} 추출 오류: {e}")
                continue

        # 추출 완료 후 자동 재배정 실행
        print("[백그라운드] 리뷰 자동 재배정 시작")
        auto_redistribute_reviews(db)
        print("[백그라운드] 리뷰 자동 재배정 완료")

    finally:
        db.close()

# 수동 리뷰 재배정 API
@app.post("/api/admin/reviews/redistribute")
async def redistribute_reviews(
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """초과 리뷰를 다음 주문으로 수동 재배정"""
    try:
        auto_redistribute_reviews(db)

        # 재배정 결과 확인
        orders = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.completed_count > ReceiptWorkOrder.total_count
        ).all()

        excess_count = sum(order.completed_count - order.total_count for order in orders)

        return {
            "success": True,
            "message": "리뷰 재배정 완료",
            "excess_reviews": excess_count,
            "affected_orders": len(orders)
        }
    except Exception as e:
        return {"success": False, "message": f"재배정 실패: {str(e)}"}

# 리뷰 내용 추출 API (비동기)
@app.post("/api/admin/reviews/extract-all")
async def extract_all_reviews(
    background_tasks: BackgroundTasks,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """모든 미추출 리뷰 내용 일괄 추출 (백그라운드)"""
    try:
        # 내용이 없거나 "내용 추출 대기중"인 리뷰들 조회
        reviews = db.query(Review).filter(
            or_(
                Review.content == "내용 추출 대기중",
                Review.content == None,
                Review.content == ""
            )
        ).all()

        review_ids = [r.id for r in reviews]
        total_count = len(review_ids)

        if total_count == 0:
            return {
                "success": True,
                "message": "추출할 리뷰가 없습니다.",
                "total_count": 0
            }

        # 백그라운드 태스크로 실행
        background_tasks.add_task(extract_reviews_background, review_ids)

        return {
            "success": True,
            "message": f"{total_count}개 리뷰 추출을 시작했습니다. 백그라운드에서 진행됩니다.",
            "total_count": total_count
        }

    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/admin/reviews/{review_id}/extract")
async def extract_single_review(
    review_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """개별 리뷰 내용 재추출"""
    try:
        from real_review_extractor import get_extractor
        extractor = get_extractor()

        review = db.query(Review).filter(Review.id == review_id).first()
        if not review:
            return {"success": False, "message": "리뷰를 찾을 수 없습니다."}

        # 업체명 가져오기
        order = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.id == review.order_id
        ).first()

        shop_name = order.business_name if order else None

        # URL에서 리뷰 내용 추출
        review_text, receipt_date, metadata = extractor.extract_review(
            review.review_url,
            shop_name
        )

        if review_text and "오류" not in review_text and "찾을 수 없습니다" not in review_text:
            review.content = review_text
            db.commit()
            return {
                "success": True,
                "message": "리뷰 내용 추출 성공",
                "content": review_text
            }
        else:
            return {
                "success": False,
                "message": f"추출 실패: {review_text}"
            }

    except Exception as e:
        return {"success": False, "message": str(e)}

# 고객사 주문 삭제 (자신의 주문만)
@app.delete("/api/client/orders/{order_id}")
async def delete_client_order(
    order_id: int,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """고객사가 자신의 주문 삭제"""
    try:
        order = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.id == order_id,
            ReceiptWorkOrder.client_id == user.id
        ).first()

        if not order:
            return JSONResponse({"success": False, "message": "주문을 찾을 수 없습니다."}, status_code=404)

        # 대기중 상태만 삭제 가능
        if order.status != 'pending':
            return JSONResponse({"success": False, "message": "대기중 상태의 주문만 삭제할 수 있습니다."}, status_code=400)

        # 관련 리뷰 삭제
        db.query(Review).filter(Review.order_id == order_id).delete()

        # 주문 삭제
        db.delete(order)
        db.commit()

        return {"success": True, "message": "주문이 삭제되었습니다."}

    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# 관리자 주문 삭제
@app.delete("/api/admin/orders/{order_id}")
async def delete_admin_order(
    order_id: int,
    user = Depends(require_super_admin),
    db: Session = Depends(get_db)
):
    """관리자가 주문 삭제"""
    try:
        order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.id == order_id).first()

        if not order:
            return JSONResponse({"success": False, "message": "주문을 찾을 수 없습니다."}, status_code=404)

        # 관련 리뷰 삭제
        review_count = db.query(Review).filter(Review.order_id == order_id).count()
        db.query(Review).filter(Review.order_id == order_id).delete()

        # 주문 삭제
        db.delete(order)
        db.commit()

        return {"success": True, "message": f"주문과 관련 리뷰 {review_count}개가 삭제되었습니다."}

    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# 리뷰 조회 API
@app.get("/api/receipt/reviews/{order_id}")
async def get_order_reviews(
    order_id: int,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """주문의 리뷰 목록 조회"""
    try:
        # 주문 확인
        order = db.query(ReceiptWorkOrder).filter(
            ReceiptWorkOrder.id == order_id,
            ReceiptWorkOrder.client_id == user.id
        ).first()

        if not order:
            return JSONResponse({"success": False, "message": "주문을 찾을 수 없습니다."}, status_code=404)

        # 리뷰 조회 - 추출 실패한 리뷰는 제외
        reviews = db.query(Review).filter(
            Review.order_id == order_id,
            ~Review.content.like("추출 실패%")
        ).all()

        review_data = []
        for review in reviews:
            review_data.append({
                "id": review.id,
                "receipt_date": review.receipt_date,
                "url": review.review_url,
                "content": review.content or "내용 없음"
            })

        return {"success": True, "reviews": review_data}

    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# ============================================
# 영수증 생성기 (관리자 전용)
# ============================================

@app.get("/uploads/orders/{filename}")
async def get_order_image(
    filename: str,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """주문 첨부 이미지 조회 (업체명으로 파일명 변경)"""
    file_path = os.path.join("uploads/orders", filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="이미지를 찾을 수 없습니다")

    # 파일명에서 주문번호 추출 (예: WO20250114001_1.jpg)
    order_no = filename.split('_')[0]

    # DB에서 주문 정보 조회하여 업체명 가져오기
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.order_no == order_no).first()

    if order:
        # 파일 확장자 및 번호 추출
        _, ext = os.path.splitext(filename)
        file_number = filename.split('_')[1].split('.')[0] if '_' in filename else '1'

        # 다운로드 파일명 생성
        download_name = f"{order.business_name}_영수증_{file_number}{ext}"
    else:
        download_name = filename

    return FileResponse(file_path, filename=download_name)

@app.get("/static/{filename}")
async def get_static_file(filename: str):
    """정적 파일 제공 (템플릿 등)"""
    file_path = os.path.join("static", filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")
    return FileResponse(file_path)

@app.get("/uploads/review_assets/{filename}")
async def get_review_asset(
    filename: str,
    user = Depends(require_login),
    db: Session = Depends(get_db)
):
    """리뷰 자료 파일 다운로드 (업체명으로 파일명 변경)"""
    file_path = os.path.join("uploads/review_assets", filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

    # 파일명에서 주문번호 추출 (예: WO20250114001_review.xlsx)
    order_no = filename.split('_')[0]

    # DB에서 주문 정보 조회하여 업체명 가져오기
    order = db.query(ReceiptWorkOrder).filter(ReceiptWorkOrder.order_no == order_no).first()

    if order:
        # 파일 확장자 추출
        _, ext = os.path.splitext(filename)

        # 파일 타입에 따라 다운로드 파일명 생성
        if '_review' in filename and not '_photos' in filename:
            download_name = f"{order.business_name}_리뷰멘트{ext}"
        elif '_photos' in filename:
            download_name = f"{order.business_name}_리뷰사진{ext}"
        else:
            download_name = f"{order.business_name}{ext}"
    else:
        download_name = filename

    return FileResponse(file_path, filename=download_name)

@app.get("/admin/receipt-generator")
async def receipt_generator_page(
    request: Request,
    user = Depends(require_super_admin)
):
    """영수증 생성기 페이지 (관리자 전용)"""
    return templates.TemplateResponse("receipt_generator.html", {
        "request": request,
        "user": user
    })

@app.post("/api/admin/receipt/fetch-menu")
async def fetch_naver_menu_api(
    request: Request,
    db: Session = Depends(get_db)
):
    """네이버 플레이스에서 메뉴 가져오기"""
    try:
        # 수동 인증 처리 (JSON 응답 보장)
        user = get_current_user(request, db)
        if not user:
            return JSONResponse({
                "success": False,
                "message": "로그인이 필요합니다."
            }, status_code=401)

        if user.role != "super_admin":
            return JSONResponse({
                "success": False,
                "message": "관리자 권한이 필요합니다."
            }, status_code=403)

        from receipt_generator.naver_scraper import get_naver_place_menu, format_menu_for_textarea

        body = await request.json()
        url = body.get('url')

        if not url:
            return JSONResponse({
                "success": False,
                "message": "URL을 입력해주세요."
            }, status_code=400)

        # 메뉴 가져오기
        menu_list = get_naver_place_menu(url)

        if not menu_list:
            return JSONResponse({
                "success": False,
                "message": "메뉴를 찾을 수 없습니다. URL을 확인해주세요."
            }, status_code=400)

        # 7글자 필터링 적용
        from receipt_generator.receipt_generator import smart_filter_menu
        filtered_menu = []
        for menu_name, price in menu_list:
            filtered_name = smart_filter_menu(menu_name, max_length=7)
            if filtered_name:
                filtered_menu.append((filtered_name, price))

        if not filtered_menu:
            return JSONResponse({
                "success": False,
                "message": "7글자 이하로 필터링한 결과 메뉴가 없습니다. 직접 입력해주세요."
            }, status_code=400)

        # 메뉴 텍스트로 변환
        menu_text = format_menu_for_textarea(filtered_menu)

        return JSONResponse({
            "success": True,
            "menu_text": menu_text,
            "menu_count": len(filtered_menu),
            "original_count": len(menu_list)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({
            "success": False,
            "message": f"메뉴 가져오기 실패: {str(e)}"
        }, status_code=500)

@app.post("/api/admin/receipt/generate")
async def generate_receipt_api(
    request: Request,
    store_name: str = Form(...),
    biz_num: str = Form(...),
    owner_name: str = Form(...),
    tel: str = Form(...),
    address: str = Form(...),
    menu_text: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    daily_count: int = Form(...),
    start_hour: int = Form(9),
    end_hour: int = Form(22),
    apply_filter: bool = Form(False),
    db: Session = Depends(get_db)
):
    """영수증 생성 API"""
    try:
        # 수동 인증 처리 (JSON 응답 보장)
        user = get_current_user(request, db)
        if not user:
            return JSONResponse({
                "success": False,
                "message": "로그인이 필요합니다."
            }, status_code=401)

        if user.role != "super_admin":
            return JSONResponse({
                "success": False,
                "message": "관리자 권한이 필요합니다."
            }, status_code=403)
        from receipt_generator.receipt_generator import generate_receipts_batch_web, parse_menu_input
        from datetime import datetime
        import zipfile
        import io

        # 날짜 파싱
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")

        # 상점 정보
        store_info = {
            "상호명": store_name,
            "사업자번호": biz_num,
            "대표자명": owner_name,
            "전화번호": tel,
            "주소": address
        }

        # 메뉴 파싱
        menu_pool = parse_menu_input(menu_text, apply_filter=apply_filter)

        if not menu_pool:
            return JSONResponse({
                "success": False,
                "message": "메뉴를 파싱할 수 없습니다. 형식을 확인해주세요."
            }, status_code=400)

        # 영수증 생성
        results = generate_receipts_batch_web(
            store_info, menu_pool, start, end, daily_count, start_hour, end_hour
        )

        # ZIP 파일 생성
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for img_io, fname in results:
                zip_file.writestr(fname, img_io.getvalue())

        zip_buffer.seek(0)

        # ZIP 파일 반환 (한글 파일명 인코딩)
        from urllib.parse import quote
        filename = f"영수증_{store_name}_{start_date}_{end_date}.zip"
        encoded_filename = quote(filename)

        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({
            "success": False,
            "message": f"영수증 생성 실패: {str(e)}"
        }, status_code=500)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)